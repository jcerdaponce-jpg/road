
import numpy as np
import math
from openpyxl import load_workbook
from collections import defaultdict
import pandas as pd
import os
from itertools import combinations

# =========================
# Opción 8: control de plots
# =========================
# True = NO dibujar (no importar Matplotlib)
# False = SÍ dibujar (importa Matplotlib y ejecuta los bloques de plotting)
NO_PLOT = True

# =========================
# INPUTS / Parámetros globales
# =========================
def main_set_medium_voltage():
    # Technical inputs (se piden por input() tal como en tu script original)
    n_clusters_max = 5
    wtg_unit_power = float(input('Unit Power of the WTG? (MW): ')) * 10**6
    medium_voltage_level = float(input('Medium Voltage Level? (kV): ')) * 10**3
    distancia_radio_desde_SET = float(input('Maximum distance from SET (meters): '))
    busbar_limitation = float(input('Busbar current limitation ? (A) (usually 2500 A): '))
    max_trafo_power = float(input('Maximum transformer power (MW) (usually 300MW): ')) * 10**6
    safety_factor_length = 1.08  # porcentaje aplicado a la longitud en planta, normalmente 8%

    # CAPEX calculation inputs (placeholder; tu comentario indica que luego vendrán de DB)
    cost_MV_Collector_system_300_MVA = 1   # €/MV
    cost_MV_Collector_system_10_5_km = 1   # €/MV
    cost_transformer_300_MVA = 4 * 10**6   # €
    cost_busbar_2500_A = 5 * 10**3         # €
    cost_SET = 15 * 10**6                  # €

    # Other inputs
    coordenadas_wtg_excel = "COORDENADAS_ZONA_2.xlsx"
    wb = load_workbook(coordenadas_wtg_excel, read_only=True)  # Excel debe estar en la misma carpeta y CERRADO
    sheet = wb.active

    # Leer todas las filas del excel y almacenarlas como listas dentro de wtg_info
    wtg_info = []
    for fila in sheet.iter_rows(values_only=True):
        wtg_info.append(list(fila))

    # Número de WTGs
    n_wtg = len(wtg_info)
    # Potencia total del parque (MW -> W)
    pot_total_windfarm = n_wtg * wtg_unit_power

    # Datos extremos (no usados directamente ahora)
    nombre_wtg = [fila[0] for fila in wtg_info]
    mayor_coord_X = max(fila[1] for fila in wtg_info)
    mayor_coord_Y = max(fila[2] for fila in wtg_info)
    menor_coord_X = min(fila[1] for fila in wtg_info)
    menor_coord_Y = min(fila[2] for fila in wtg_info)

    # ----- FUNCIONES AUXILIARES (usan variables de cierre) -----

    def asignar_wtgs_a_sets(sets_coord, wtg_info_local, usar_indice_desde_1=True):
        resultados = {}
        wtg_en_sets = defaultdict(list)
        asignaciones = {}

        for fila in wtg_info_local:
            id_wtg, x_wtg, y_wtg = fila[0], fila[1], fila[2]
            min_dist = float('inf')
            set_idx_min = None
            distancias = []
            for idx, (x_set, y_set) in enumerate(sets_coord):
                dist = math.hypot(x_wtg - x_set, y_wtg - y_set)
                distancias.append(dist)
                if dist < min_dist:
                    min_dist = dist
                    set_idx_min = idx
            dist_min_km = min(distancias) / 1000.0
            dist_max_km = max(distancias) / 1000.0
            resultados[id_wtg] = {"dist_min": dist_min_km, "dist_max": dist_max_km}
            indice_reportado = set_idx_min + 1 if usar_indice_desde_1 else set_idx_min
            wtg_en_sets[indice_reportado].append(id_wtg)
            asignaciones[id_wtg] = {"set": indice_reportado, "dist": min_dist}
        return wtg_en_sets, asignaciones, resultados

    def resumen_por_set_completo(sets_coord, wtg_info_local, wtg_en_sets, usar_vecino_mas_cercano=True):
        wtg_dict = {fila[0]: (fila[1], fila[2]) for fila in wtg_info_local}
        resumen = []
        for set_idx, lista_wtgs in wtg_en_sets.items():
            if not lista_wtgs:
                resumen.append({
                    "SET": set_idx,
                    "WTGs": 0,
                    "Dist_Min": None,
                    "Dist_Max": None,
                    "Dist_Media_WTGs": None
                })
                continue
            x_set, y_set = sets_coord[set_idx - 1]
            distancias_set = [math.hypot(wtg_dict[w][0] - x_set, wtg_dict[w][1] - y_set) for w in lista_wtgs]
            dist_min = min(distancias_set) / 1000.0
            dist_max = max(distancias_set) / 1000.0

            if len(lista_wtgs) < 2:
                dist_media_wtgs = None
            else:
                if usar_vecino_mas_cercano:
                    distancias_minimas = []
                    for wtg_id in lista_wtgs:
                        distancias_otras = [
                            math.hypot(wtg_dict[wtg_id][0] - wtg_dict[otro][0],
                                       wtg_dict[wtg_id][1] - wtg_dict[otro][1])
                            for otro in lista_wtgs if otro != wtg_id
                        ]
                        distancias_minimas.append(min(distancias_otras))
                    dist_media_wtgs = (sum(distancias_minimas) / len(distancias_minimas)) / 1000.0
                else:
                    distancias_parejas = [
                        math.hypot(wtg_dict[a][0] - wtg_dict[b][0], wtg_dict[a][1] - wtg_dict[b][1])
                        for a, b in combinations(lista_wtgs, 2)
                    ]
                    dist_media_wtgs = (sum(distancias_parejas) / len(distancias_parejas)) / 1000.0

            resumen.append({
                "SET": set_idx,
                "WTGs": len(lista_wtgs),
                "Dist_Min": dist_min,
                "Dist_Max": dist_max,
                "Dist_Media_WTGs": dist_media_wtgs
            })
        return resumen

    # --- Cálculo de pérdidas por SET (usa tu lógica original y parámetros globales) ---
    def calculo_perdidas(resumen_sets_funcion, safety_factor_length_local):
        # Aceptar dict {"resumen_sets": [...]} o lista
        if isinstance(resumen_sets_funcion, dict) and "resumen_sets" in resumen_sets_funcion:
            resumen_sets = resumen_sets_funcion["resumen_sets"]
        else:
            resumen_sets = resumen_sets_funcion

        # Corrientes por circuito (PF=0.95)
        current_1WTG = (((wtg_unit_power / 1000) / 0.95) / (math.sqrt(3) * medium_voltage_level / 1000))
        current_2WTG = ((((2 * wtg_unit_power) / 1000) / 0.95) / (math.sqrt(3) * medium_voltage_level / 1000))
        current_3WTG = ((((3 * wtg_unit_power) / 1000) / 0.95) / (math.sqrt(3) * medium_voltage_level / 1000))

        # Resistencias (Ohm/km)
        resist_120mm2 = 0.325
        resist_300mm2 = 0.130
        resist_630mm2 = 0.063

        # Listas para devolver valores por SET
        perdidas_3WTG_por_set = []
        perdidas_2WTG_por_set = []
        perdidas_1WTG_por_set = []
        supply_120mm2_3WTGs = []
        supply_300mm2_3WTGs = []
        supply_630mm2_3WTGs = []
        supply_120mm2_2WTGs = []
        supply_300mm2_2WTGs = []
        supply_630mm2_1WTGs = []

        for r in resumen_sets:
            dist_media = r["Dist_Media_WTGs"]
            dist_min = r["Dist_Min"]
            dist_max = r["Dist_Max"]
            if dist_media is None or dist_min is None or dist_max is None:
                perdidas_3WTG_por_set.append(0.0)
                perdidas_2WTG_por_set.append(0.0)
                perdidas_1WTG_por_set.append(0.0)
                supply_120mm2_3WTGs.append(0.0)
                supply_300mm2_3WTGs.append(0.0)
                supply_630mm2_3WTGs.append(0.0)
                supply_120mm2_2WTGs.append(0.0)
                supply_300mm2_2WTGs.append(0.0)
                supply_630mm2_1WTGs.append(0.0)
                continue

            # Aplica factor de seguridad
            d_btw = dist_media * safety_factor_length_local
            d_mean = ((dist_max + dist_min) / 2) * safety_factor_length_local

            # Circuitos de 3 WTGs
            p_1_3 = 3 * (current_1WTG ** 2) * (d_btw * resist_120mm2)
            s_120_3 = d_btw * 3
            p_2_3 = 3 * (current_2WTG ** 2) * (d_btw * resist_300mm2)
            s_300_3 = d_btw * 3
            p_3_3 = 3 * (current_3WTG ** 2) * (d_mean * resist_630mm2)
            s_630_3 = d_mean * 3
            perdidas_3WTG_por_set.append(p_1_3 + p_2_3 + p_3_3)
            supply_120mm2_3WTGs.append(s_120_3)
            supply_300mm2_3WTGs.append(s_300_3)
            supply_630mm2_3WTGs.append(s_630_3)

            # Circuitos de 2 WTGs
            p_1_2 = 3 * (current_1WTG ** 2) * (d_btw * resist_120mm2)
            s_120_2 = d_btw * 3
            p_2_2 = 3 * (current_2WTG ** 2) * (d_mean * resist_630mm2)
            s_300_2 = d_mean * 3  # Mantengo tu lógica
            perdidas_2WTG_por_set.append(p_1_2 + p_2_2)
            supply_120mm2_2WTGs.append(s_120_2)
            supply_300mm2_2WTGs.append(s_300_2)

            # Circuitos de 1 WTG
            p_1_1 = 3 * (current_1WTG ** 2) * (d_mean * resist_630mm2)
            s_630_1 = d_mean * 3
            perdidas_1WTG_por_set.append(p_1_1)
            supply_630mm2_1WTGs.append(s_630_1)

        return (
            perdidas_3WTG_por_set,
            perdidas_2WTG_por_set,
            perdidas_1WTG_por_set,
            supply_120mm2_3WTGs,
            supply_300mm2_3WTGs,
            supply_630mm2_3WTGs,
            supply_120mm2_2WTGs,
            supply_300mm2_2WTGs,
            supply_630mm2_1WTGs
        )

    # --- Procesar excels de resultados por configuración / opción ---
    def procesar_excels(
        wtg_opcion1, wtg_opcion2, wtg_opcion3, wtg_opcion4,
        carpeta_resultados, resumen_sets_funcion,
        archivo_final="resultados_circuitos.xlsx",
        archivo_perdidas="Pérdidas Totales_in kW.xlsx",
        archivos_filtrar=None
    ):
        # Aceptar dict {"resumen_sets": [...]} o lista
        if isinstance(resumen_sets_funcion, dict) and "resumen_sets" in resumen_sets_funcion:
            resumen_sets_funcion = resumen_sets_funcion["resumen_sets"]
        if not isinstance(resumen_sets_funcion, list):
            raise ValueError("Se esperaba una lista de diccionarios en resumen_sets_funcion.")

        (
            p3_list, p2_list, p1_list,
            s120_3_list, s300_3_list, s630_3_list,
            s120_2_list, s300_2_list, s630_1_list
        ) = calculo_perdidas(resumen_sets_funcion, safety_factor_length)

        # Selección de archivos a procesar
        if archivos_filtrar and len(archivos_filtrar) > 0:
            archivos = archivos_filtrar
        else:
            archivos = [f for f in os.listdir(carpeta_resultados) if f.endswith('.xlsx')]

        # Acumuladores globales
        lista_cantidad = []
        lista_circuitos_3 = []
        lista_circuitos_2 = []
        lista_cantidad_120_mm2 = []
        lista_cantidad_300_mm2 = []
        lista_cantidad_630_mm2 = []
        dataframes = []
        solo_perdidas_pruebas = []

        for archivo in archivos:
            ruta_completa = os.path.join(carpeta_resultados, archivo)
            try:
                df = pd.read_excel(ruta_completa)
            except FileNotFoundError:
                ruta_alt = os.path.join(carpeta_resultados, "WTGs fuera de radio", archivo)
                df = pd.read_excel(ruta_alt)
            if 'Cantidad por SET' not in df.columns:
                print(f"⚠ El archivo '{archivo}' no tiene la columna 'Cantidad por SET'. Se omite.")
                continue

            cantidad = df['Cantidad por SET']

            # Coherencia: nº filas vs nº SETs en pérdidas
            if len(cantidad) != len(p3_list):
                raise ValueError(
                    f"Desajuste de longitudes para '{archivo}': "
                    f"{len(cantidad)} filas (SETs en Excel) vs {len(p3_list)} SETs en pérdidas.\n"
                    f"Asegúrate de pasar a 'archivos_filtrar' el archivo correcto para esta configuración."
                )

            # Derivar nº circuitos de 3 y 2 WTGs (tu regla de 0.333 / 0.667)
            n_circuitos_3_wtg = cantidad / 3
            n_circuitos_2_wtg = pd.Series([0] * len(cantidad))
            parte_decimal = n_circuitos_3_wtg % 1
            parte_decimal_redondeada = parte_decimal.round(3)

            for i in range(len(cantidad)):
                if parte_decimal_redondeada[i] == 0.333:
                    n_circuitos_3_wtg[i] = (cantidad[i] - 4) / 3
                    n_circuitos_2_wtg[i] = 2
                elif parte_decimal_redondeada[i] == 0.667:
                    n_circuitos_3_wtg[i] = (cantidad[i] - 2) / 3
                    n_circuitos_2_wtg[i] = 1
                else:
                    n_circuitos_3_wtg[i] = int(n_circuitos_3_wtg[i])
                    n_circuitos_2_wtg[i] = 0

            # Pérdidas y supply por SET
            perdidas_totales_SET = []
            supply_120_total = []
            supply_300_total = []
            supply_630_total = []

            for i in range(len(cantidad)):
                perdidas_3 = n_circuitos_3_wtg[i] * p3_list[i]
                perdidas_2 = n_circuitos_2_wtg[i] * p2_list[i]
                perdidas_kW = (perdidas_3 + perdidas_2) / 1000.0
                perdidas_totales_SET.append(perdidas_kW)

                s120 = (n_circuitos_3_wtg[i] * s120_3_list[i]) * 3 + (n_circuitos_2_wtg[i] * s120_2_list[i]) * 3
                s300 = (n_circuitos_3_wtg[i] * s300_3_list[i]) * 3 + (n_circuitos_2_wtg[i] * s300_2_list[i]) * 3
                s630 = (n_circuitos_3_wtg[i] * s630_3_list[i]) * 3 + (n_circuitos_2_wtg[i] * s630_1_list[i]) * 3
                supply_120_total.append(s120)
                supply_300_total.append(s300)
                supply_630_total.append(s630)

            # Totales globales
            lista_cantidad.append(cantidad.sum())
            lista_circuitos_3.append(n_circuitos_3_wtg.sum())
            lista_circuitos_2.append(n_circuitos_2_wtg.sum())
            lista_cantidad_120_mm2.append(sum(supply_120_total))
            lista_cantidad_300_mm2.append(sum(supply_300_total))
            lista_cantidad_630_mm2.append(sum(supply_630_total))

            # DataFrame por archivo
            nombre_base = os.path.splitext(archivo)[0]
            df_resultado = pd.DataFrame({
                nombre_base + '_Cir_3WTG': n_circuitos_3_wtg,
                nombre_base + '_Cir_2WTG': n_circuitos_2_wtg,
                nombre_base + '_Total_Cantidad': cantidad,
                nombre_base + '_Pérdidas (kW)': perdidas_totales_SET,
                nombre_base + '_Supply 120mm2 (km)': supply_120_total,
                nombre_base + '_Supply 300mm2 (km)': supply_300_total,
                nombre_base + '_Supply 630mm2 (km)': supply_630_total,
            })
            dataframes.append(df_resultado)

            # Solo pérdidas totales (por archivo)
            total_perdidas = sum(perdidas_totales_SET)
            df_solo_perdidas = pd.DataFrame({nombre_base + '_PÉRDIDAS TOTALES': [total_perdidas]})
            solo_perdidas_pruebas.append(df_solo_perdidas)

        # Combinar y exportar
        df_final = pd.concat(dataframes, axis=1) if dataframes else pd.DataFrame()
        df_solo_perdidas_final = pd.concat(solo_perdidas_pruebas, axis=1).T.reset_index() if solo_perdidas_pruebas else pd.DataFrame(columns=['Configuración','Pérdidas en kW'])
        if not df_solo_perdidas_final.empty:
            df_solo_perdidas_final.columns = ['Configuración', 'Pérdidas en kW']
            df_solo_perdidas_final['Pérdidas en %'] = ((df_solo_perdidas_final['Pérdidas en kW'] * 1000) / pot_total_windfarm) * 100
            df_solo_perdidas_final['Circuitos de 3 WTGs'] = lista_circuitos_3
            df_solo_perdidas_final['Circuitos de 2 WTGs'] = lista_circuitos_2
            df_solo_perdidas_final['Número Total Circuitos'] = df_solo_perdidas_final['Circuitos de 3 WTGs'] + df_solo_perdidas_final['Circuitos de 2 WTGs']
            df_solo_perdidas_final['Supply 120mm2 cable (km)'] = lista_cantidad_120_mm2
            df_solo_perdidas_final['Supply 300mm2 cable (km)'] = lista_cantidad_300_mm2
            df_solo_perdidas_final['Supply 630mm2 cable (km)'] = lista_cantidad_630_mm2
            df_solo_perdidas_final['Supply earthing cable (km)'] = (
                df_solo_perdidas_final['Supply 120mm2 cable (km)'] +
                df_solo_perdidas_final['Supply 300mm2 cable (km)'] +
                df_solo_perdidas_final['Supply 630mm2 cable (km)']
            )

        if not df_final.empty:
            df_final.to_excel(archivo_final, index=False)
        if not df_solo_perdidas_final.empty:
            df_solo_perdidas_final.to_excel(archivo_perdidas, index=False)

        # Estructura para recopilación
        resumen_config = []
        if not df_solo_perdidas_final.empty:
            for _, fila in df_solo_perdidas_final.iterrows():
                resumen_config.append({
                    "Configuración": fila["Configuración"],
                    "Pérdidas en kW": fila["Pérdidas en kW"],
                    "Pérdidas en %": fila["Pérdidas en %"],
                    "Circuitos 3 WTGs": fila["Circuitos de 3 WTGs"],
                    "Circuitos 2 WTGs": fila["Circuitos de 2 WTGs"],
                    "Nº Total Circuitos": fila["Número Total Circuitos"],
                    "Supply 120mm2 (km)": fila["Supply 120mm2 cable (km)"],
                    "Supply 300mm2 (km)": fila["Supply 300mm2 cable (km)"],
                    "Supply 630mm2 (km)": fila["Supply 630mm2 cable (km)"],
                    "Supply earthing (km)": fila["Supply earthing cable (km)"]
                })

        detalle_por_set = df_final.copy() if not df_final.empty else pd.DataFrame()
        return {
            "archivo_resultados": archivo_final,
            "archivo_perdidas": archivo_perdidas,
            "resumen_por_config": resumen_config,
            "detalle_por_set": detalle_por_set
        }

    # --- Cálculos comunes (después de asignación y resumen por SET) ---
    def calculos_comunes(sets_coord, criterio_distancia, wtgs_por_set_filtrada, opcion, n_clusters_max_local, wtg_fuera_radio):
        wtg_en_sets, asignaciones, resultados = asignar_wtgs_a_sets(sets_coord, wtg_info, usar_indice_desde_1=True)
        resumen_sets_funcion_local = resumen_por_set_completo(sets_coord, wtg_info, wtg_en_sets, usar_vecino_mas_cercano=True)

        # Alinear longitudes si hay discrepancias
        if len(sets_coord) != len(wtgs_por_set_filtrada):
            mask_no_vacio = [len(sublista) > 0 for sublista in wtgs_por_set_filtrada]
            if all(mask_no_vacio):
                sets_coord = np.array(sets_coord, dtype=float)[:len(wtgs_por_set_filtrada)]
            else:
                sets_coord = np.array(sets_coord, dtype=float)[mask_no_vacio]
                wtgs_por_set_filtrada = [s for s in wtgs_por_set_filtrada if len(s) > 0]

        n_sets = len(sets_coord)

        # Conteo por SET
        conteo_por_set = [len(lista_wtgs) for lista_wtgs in wtgs_por_set_filtrada]
        lista_cantidad = []
        for i, cantidad in enumerate(conteo_por_set):
            lista_cantidad.append(cantidad)

        pot_per_SET_MW = [valor * wtg_unit_power / (10**6) for valor in lista_cantidad]

        # Busbars por SET
        number_busbar = [(((valor * 10**6) / (math.sqrt(3) * medium_voltage_level)) / busbar_limitation) for valor in pot_per_SET_MW]
        number_busbar_redondeados_arriba = [math.ceil(valor) for valor in number_busbar]
        n_totales_busbar = sum(number_busbar_redondeados_arriba)

        # Transformadores por SET
        number_transformers = [((valor * 10**6) / max_trafo_power) for valor in pot_per_SET_MW]
        number_transformers_redondeados_arriba = [math.ceil(valor) for valor in number_transformers]
        n_totales_transformers = sum(number_transformers_redondeados_arriba)

        # CAPEX
        if criterio_distancia == 1:
            CAPEX = cost_MV_Collector_system_10_5_km * n_wtg * wtg_unit_power + \
                    n_totales_busbar * cost_busbar_2500_A + \
                    n_totales_transformers * cost_transformer_300_MVA + \
                    n_sets * cost_SET
        else:
            CAPEX = cost_MV_Collector_system_300_MVA * n_wtg * wtg_unit_power + \
                    n_totales_busbar * cost_busbar_2500_A + \
                    n_totales_transformers * cost_transformer_300_MVA + \
                    n_sets * cost_SET

        # Excel de resumen (coordenadas y conteos)
        sets_coord_str = [f"[{x[0]:.6f}, {x[1]:.6f}]" for x in sets_coord]
        capex_column = ["" for _ in sets_coord_str]
        capex_formateado = f"{CAPEX:,.2f}"


main_set_medium_voltage()