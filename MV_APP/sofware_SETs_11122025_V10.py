import numpy as np
import matplotlib.pyplot as plt
from sklearn.cluster import KMeans
import math
from openpyxl import load_workbook
from collections import defaultdict
import pandas as pd
import numpy as np
import os
import pandas as pd
import matplotlib.patches as patches
from matplotlib.patches import Circle, Patch
from itertools import combinations


############INPUTS#########################################################################################################################################
def main_set_medium_voltage():
    #Technical inputs
    n_clusters_max=1
    wtg_unit_power=float(input('Unit Power of the WTG? (MW): '))*10**6
    medium_voltage_level=float(input('Medium Voltage Level? (kV): '))*10**3
    distancia_radio_desde_SET=float(input('Maximum distance from SET (meters): '))
    busbar_limitation=float(input('Busbar current limitation ? (A) (usually 2500 A): '))
    max_trafo_power=float(input('Maximum transformer power (MW) (usually 300MW): '))*10**6

    safety_factor_length=1.08 #Porcentaje aplicado a la longitud en planta, normalmente 8%



    #CAPEX calculation inputs (valores absolutamente inventados por ahora). La idea es que los coja de la base de datos
    cost_MV_Collector_system_300_MVA = 1 #€/MV
    cost_MV_Collector_system_10_5_km = 1 #€/MV
    cost_transformer_300_MVA = 4*10**6 #€
    cost_busbar_2500_A = 5*10**3 #€
    cost_SET = 15*10**6 #€ (WARNING: ¿NO TENDRÍA MÁS SENTIDO PONER UN COSTE DE SET POR MEGAWATIO?. NO CREO QUE SEA LO MISMO UNA SET DE 300MW QUE UNA DE 1000MW)

    #Other inputs
    coordenadas_wtg_excel="COORDENADAS_ZONA_2.xlsx"
    # coordenadas_wtg_excel="WTGS_COORDINATES_400_CdM.xlsx"
    # coordenadas_wtg_excel="PRONGHORN_DEF.xlsx"

    # Estos inputs todavía no se usan para nada por ahora.
    # wtg_model=str(input('WTG model?: '))
    # tower_type=str(input('Tower type (S/C/H): '))
    # tower_height=str(input('Tower height?: '))
    ##########


    ###################################################################################################################################
    #1º se lee el excel donde se encuentras las coordenadas de toods los WTGs
    #El formato debe de ser: 1ºcolumna: Nombre WTG, 2ºcolumna: Coordenada X, 3ºcolumna: Coordenada Y
    # wb=load_workbook("WTGS_COORDINATES_400_CdM.xlsx", read_only=True)#El excell debe estar en la misma carpeta que este código y CERRADO
    wb=load_workbook(coordenadas_wtg_excel, read_only=True)#El excell debe estar en la misma carpeta que este código y CERRADO

    sheet=wb.active

    #Leer todas las filas del excel y almacenarlas como listas dentro de una lista global llamada wtg_info
    wtg_info=[]
    for fila in sheet.iter_rows(values_only=True):
        wtg_info.append(list(fila))

    #Calcula el número de WTGs que hay
    n_wtg=len(wtg_info)


    #Calcula la potencia total del parque
    pot_total_windfarm = n_wtg*wtg_unit_power


    #Saca la mayores y menores coordenadas para ver los 'límites' físicos del parque (no se usa por ahora)
    nombre_wtg= [fila[0] for fila in wtg_info]
    mayor_coord_X = max(fila[1] for fila in wtg_info)
    mayor_coord_Y = max(fila[2] for fila in wtg_info)
    menor_coord_X = min(fila[1] for fila in wtg_info)
    menor_coord_Y = min(fila[2] for fila in wtg_info)

    #print("Mayor coord X = ", mayor_coord_X)
    # print("Mayor coord Y = ", mayor_coord_Y)
    # print("Menor coord X = ", menor_coord_X)
    # print("Menor coord Y = ", menor_coord_Y)

    ######################################################################################################################
    def asignar_wtgs_a_sets(sets_coord, wtg_info, usar_indice_desde_1=True):
        resultados = {}
        # print('sets_coord',sets_coord)
        wtg_en_sets = defaultdict(list)
        asignaciones = {}

        for fila in wtg_info:
            id_wtg, x_wtg, y_wtg = fila[0], fila[1], fila[2]

            # Encuentra el SET más cercano
            min_dist = float('inf')
            set_idx_min = None
            distancias = []
            for idx, (x_set, y_set) in enumerate(sets_coord):
                dist = math.hypot(x_wtg - x_set, y_wtg - y_set)  # equivalente a sqrt(dx^2 + dy^2)
                distancias.append(dist)
                if dist < min_dist:
                    min_dist = dist
                    set_idx_min = idx

            dist_min=min(distancias)/1000 #con este /1000 se pasan las distancias a kilómetros
            dist_max=max(distancias)/1000 #con este /1000 se pasan las distancias a kilómetros

            resultados[id_wtg] = {"dist_min": dist_min, "dist_max": dist_max}

            # Ajusta índice si quieres empezar en 1 (más habitual en reportes)
            indice_reportado = set_idx_min + 1 if usar_indice_desde_1 else set_idx_min

            # Guarda agrupación y asignación
            wtg_en_sets[indice_reportado].append(id_wtg)
            asignaciones[id_wtg] = {"set": indice_reportado, "dist": min_dist}

        return wtg_en_sets, asignaciones, resultados

    ######################################################################################################################
    def resumen_por_set_completo(sets_coord, wtg_info, wtg_en_sets, usar_vecino_mas_cercano=True):
        wtg_dict = {fila[0]: (fila[1], fila[2]) for fila in wtg_info}
        resumen = []

        for set_idx, lista_wtgs in wtg_en_sets.items():
            if not lista_wtgs:
                resumen.append({
                    "SET": set_idx,
                    "WTGs": 0,
                    "Dist_Min": None,
                    "Dist_Max": None,
                    "Dist_Media_WTGs": None
                })
                continue

            # Coordenadas del SET
            x_set, y_set = sets_coord[set_idx - 1]

            # Min y Max respecto al SET
            distancias_set = [math.hypot(wtg_dict[w][0] - x_set, wtg_dict[w][1] - y_set) for w in lista_wtgs]
            dist_min = min(distancias_set)/1000 #con este /1000 se pasan las distancias a kilómetros
            dist_max = max(distancias_set)/1000 #con este /1000 se pasan las distancias a kilómetros

            # Media entre WTGs
            if len(lista_wtgs) < 2:
                dist_media_wtgs = None
            else:
                if usar_vecino_mas_cercano:
                    # Media del vecino más cercano (más realista)
                    distancias_minimas = []
                    for wtg_id in lista_wtgs:
                        distancias_otras = [
                            math.hypot(wtg_dict[wtg_id][0] - wtg_dict[otro][0],
                                    wtg_dict[wtg_id][1] - wtg_dict[otro][1])
                            for otro in lista_wtgs if otro != wtg_id
                        ]
                        distancias_minimas.append(min(distancias_otras))
                    dist_media_wtgs = (sum(distancias_minimas) / len(distancias_minimas))/1000 #con este /1000 se pasan las distancias a kilómetros
                else:
                    # Media de todas las parejas (dispersión total)
                    distancias_parejas = [
                        math.hypot(wtg_dict[a][0] - wtg_dict[b][0],
                                wtg_dict[a][1] - wtg_dict[b][1])
                        for a, b in combinations(lista_wtgs, 2)
                    ]
                    dist_media_wtgs = (sum(distancias_parejas) / len(distancias_parejas))/1000 #con este /1000 se pasan las distancias a kilómetros

            resumen.append({
                "SET": set_idx,
                "WTGs": len(lista_wtgs),
                "Dist_Min": dist_min,
                "Dist_Max": dist_max,
                "Dist_Media_WTGs": dist_media_wtgs
            })
        return resumen
    ######################################################################################################################

    ######################PRUEBA CÁLCULO PÉRDIDAS CON LOS WTGS DENTRO DE LOS 10.5KM####################################################################################################################
    def procesar_excels(
        wtg_opcion1, wtg_opcion2, wtg_opcion3, wtg_opcion4,
        carpeta_resultados, resumen_sets_funcion,
        archivo_final="resultados_circuitos.xlsx",
        archivo_perdidas="Pérdidas Totales_in kW.xlsx",
        archivos_filtrar=None  # <-- NUEVO: lista de nombres de archivos a procesar
    ):
        """
        Procesa los archivos Excel especificados, calcula circuitos y pérdidas por SET,
        y genera dos archivos: uno con resultados completos y otro solo con pérdidas totales.
        - Si 'archivos_filtrar' es None, procesará todos los .xlsx en 'carpeta_resultados'.
        Si se pasa una lista, sólo procesará esos archivos (nombres base, sin ruta).
        """

        # Aceptar dict {"resumen_sets": [...]} o lista
        if isinstance(resumen_sets_funcion, dict) and "resumen_sets" in resumen_sets_funcion:
            resumen_sets_funcion = resumen_sets_funcion["resumen_sets"]
        if not isinstance(resumen_sets_funcion, list):
            raise ValueError("Se esperaba una lista de diccionarios en resumen_sets_funcion.")

        # Listas por SET (calculo_perdidas devuelve listas, una por SET)
        (
            p3_list, p2_list, p1_list,
            s120_3_list, s300_3_list, s630_3_list,
            s120_2_list, s300_2_list, s630_1_list
        ) = calculo_perdidas(resumen_sets_funcion, safety_factor_length)

        # --- Selección de archivos a procesar ---
        if archivos_filtrar and len(archivos_filtrar) > 0:
            archivos = archivos_filtrar  # nombres base sin ruta, p.ej. "SET calculadas... opción 1.xlsx"
        else:
            archivos = [f for f in os.listdir(carpeta_resultados) if f.endswith('.xlsx')]

        # Acumuladores globales (para df_solo_perdidas_final)
        lista_cantidad = []
        lista_circuitos_3 = []
        lista_circuitos_2 = []
        lista_cantidad_120_mm2 = []
        lista_cantidad_300_mm2 = []
        lista_cantidad_630_mm2 = []
        dataframes = []
        solo_perdidas_pruebas = []

        for archivo in archivos:
            ruta_completa = os.path.join(carpeta_resultados, archivo)
            try:
                df = pd.read_excel(ruta_completa)
            except FileNotFoundError:
                # Intenta en la subcarpeta "WTGs fuera de radio"
                ruta_alt = os.path.join(carpeta_resultados, "WTGs fuera de radio", archivo)
                df = pd.read_excel(ruta_alt)


            if 'Cantidad por SET' not in df.columns:
                print(f"⚠ El archivo '{archivo}' no tiene la columna 'Cantidad por SET'. Se omite.")
                continue

            cantidad = df['Cantidad por SET']
            # Comprobación de coherencia: nº filas del Excel vs nº SETs de las listas
            if len(cantidad) != len(p3_list):
                raise ValueError(
                    f"Desajuste de longitudes para '{archivo}': "
                    f"{len(cantidad)} filas (SETs en Excel) vs {len(p3_list)} SETs en pérdidas.\n"
                    f"Asegúrate de pasar a 'archivos_filtrar' el archivo correcto para esta configuración."
                )

            # Derivar nº circuitos 3/2 WTGs según tu lógica
            n_circuitos_3_wtg = cantidad / 3
            n_circuitos_2_wtg = pd.Series([0] * len(cantidad))
            parte_decimal = n_circuitos_3_wtg % 1
            parte_decimal_redondeada = parte_decimal.round(3)

            for i in range(len(cantidad)):
                if parte_decimal_redondeada[i] == 0.333:
                    n_circuitos_3_wtg[i] = (cantidad[i] - 4) / 3
                    n_circuitos_2_wtg[i] = 2
                elif parte_decimal_redondeada[i] == 0.667:
                    n_circuitos_3_wtg[i] = (cantidad[i] - 2) / 3
                    n_circuitos_2_wtg[i] = 1
                else:
                    n_circuitos_3_wtg[i] = int(n_circuitos_3_wtg[i])
                    n_circuitos_2_wtg[i] = 0

            # Cálculo pérdidas y cables por SET (longitudes coherentes)
            perdidas_totales_SET = []
            supply_120_total = []
            supply_300_total = []
            supply_630_total = []

            for i in range(len(cantidad)):
                perdidas_3 = n_circuitos_3_wtg[i] * p3_list[i]
                perdidas_2 = n_circuitos_2_wtg[i] * p2_list[i]
                perdidas_kW = (perdidas_3 + perdidas_2) / 1000.0
                perdidas_totales_SET.append(perdidas_kW)

                # s120 = n_circuitos_3_wtg[i] * s120_3_list[i] + n_circuitos_2_wtg[i] * s120_2_list[i]
                # s300 = n_circuitos_3_wtg[i] * s300_3_list[i] + n_circuitos_2_wtg[i] * s300_2_list[i]
                # s630 = n_circuitos_3_wtg[i] * s630_3_list[i] + n_circuitos_2_wtg[i] * s630_1_list[i]

                s120 = (n_circuitos_3_wtg[i] * s120_3_list[i])*3 + (n_circuitos_2_wtg[i] * s120_2_list[i])*3
                s300 = (n_circuitos_3_wtg[i] * s300_3_list[i])*3 + (n_circuitos_2_wtg[i] * s300_2_list[i])*3
                s630 = (n_circuitos_3_wtg[i] * s630_3_list[i])*3 + (n_circuitos_2_wtg[i] * s630_1_list[i])*3

                supply_120_total.append(s120)
                supply_300_total.append(s300)
                supply_630_total.append(s630)

            # Totales globales
            lista_cantidad.append(cantidad.sum())
            lista_circuitos_3.append(n_circuitos_3_wtg.sum())
            lista_circuitos_2.append(n_circuitos_2_wtg.sum())
            lista_cantidad_120_mm2.append(sum(supply_120_total))
            lista_cantidad_300_mm2.append(sum(supply_300_total))
            lista_cantidad_630_mm2.append(sum(supply_630_total))

            # DataFrame por archivo
            nombre_base = os.path.splitext(archivo)[0]
            df_resultado = pd.DataFrame({
                nombre_base + '_Cir_3WTG': n_circuitos_3_wtg,
                nombre_base + '_Cir_2WTG': n_circuitos_2_wtg,
                nombre_base + '_Total_Cantidad': cantidad,
                nombre_base + '_Pérdidas (kW)': perdidas_totales_SET,
                nombre_base + '_Supply 120mm2 (km)': supply_120_total,
                nombre_base + '_Supply 300mm2 (km)': supply_300_total,
                nombre_base + '_Supply 630mm2 (km)': supply_630_total,
            })
            dataframes.append(df_resultado)

            # DataFrame solo pérdidas totales (por archivo)
            total_perdidas = sum(perdidas_totales_SET)
            df_solo_perdidas = pd.DataFrame({nombre_base + '_PÉRDIDAS TOTALES': [total_perdidas]})
            solo_perdidas_pruebas.append(df_solo_perdidas)

        # Combinar y exportar
        df_final = pd.concat(dataframes, axis=1)
        df_solo_perdidas_final = pd.concat(solo_perdidas_pruebas, axis=1).T.reset_index()
        df_solo_perdidas_final.columns = ['Configuración', 'Pérdidas en kW']
        df_solo_perdidas_final['Pérdidas en %'] = ((df_solo_perdidas_final['Pérdidas en kW'] * 1000) / pot_total_windfarm) * 100
        df_solo_perdidas_final['Circuitos de 3 WTGs'] = lista_circuitos_3
        df_solo_perdidas_final['Circuitos de 2 WTGs'] = lista_circuitos_2
        df_solo_perdidas_final['Número Total Circuitos'] = df_solo_perdidas_final['Circuitos de 3 WTGs'] + df_solo_perdidas_final['Circuitos de 2 WTGs']
        df_solo_perdidas_final['Supply 120mm2 cable (km)'] = lista_cantidad_120_mm2
        df_solo_perdidas_final['Supply 300mm2 cable (km)'] = lista_cantidad_300_mm2
        df_solo_perdidas_final['Supply 630mm2 cable (km)'] = lista_cantidad_630_mm2
        df_solo_perdidas_final['Supply earthing cable (km)'] = (
            df_solo_perdidas_final['Supply 120mm2 cable (km)'] +
            df_solo_perdidas_final['Supply 300mm2 cable (km)'] +
            df_solo_perdidas_final['Supply 630mm2 cable (km)']
        )

        df_final.to_excel(archivo_final, index=False)
        df_solo_perdidas_final.to_excel(archivo_perdidas, index=False)
        # print(f"Archivo '{archivo_final}' generado exitosamente.")
        # print(f"Archivo '{archivo_perdidas}' generado exitosamente.")
        
        # === NUEVO: devolver resumen estructurado para la recopilación ===
        # 1) Totales por configuración (tomados del df_solo_perdidas_final)
        resumen_config = []
        for _, fila in df_solo_perdidas_final.iterrows():
            resumen_config.append({
                "Configuración": fila["Configuración"],
                "Pérdidas en kW": fila["Pérdidas en kW"],
                "Pérdidas en %": fila["Pérdidas en %"],
                "Circuitos 3 WTGs": fila["Circuitos de 3 WTGs"],
                "Circuitos 2 WTGs": fila["Circuitos de 2 WTGs"],
                "Nº Total Circuitos": fila["Número Total Circuitos"],
                "Supply 120mm2 (km)": fila["Supply 120mm2 cable (km)"],
                "Supply 300mm2 (km)": fila["Supply 300mm2 cable (km)"],
                "Supply 630mm2 (km)": fila["Supply 630mm2 cable (km)"],
                "Supply earthing (km)": fila["Supply earthing cable (km)"]
            })

        # 2) Detalle por SET (tomado de df_final)
        #    Armamos un resumen por SET con las columnas base que generaste
        detalle_por_set = df_final.copy()
        # La primera columna ‘*_Total_Cantidad’ te da la cantidad de WTGs por SET

        # DEVOLVER RESUMEN
        return {
            "archivo_resultados": archivo_final,
            "archivo_perdidas": archivo_perdidas,
            "resumen_por_config": resumen_config,  # lista de dicts (por configuración/archivo)
            "detalle_por_set": detalle_por_set     # DataFrame con columnas por SET
        }




    #############################################################################################################################################3333

    def calculo_perdidas(resumen_sets_funcion, safety_factor_length):
        # Si viene como dict con clave "resumen_sets", extrae la lista
        if isinstance(resumen_sets_funcion, dict) and "resumen_sets" in resumen_sets_funcion:
            resumen_sets = resumen_sets_funcion["resumen_sets"]
        else:
            resumen_sets = resumen_sets_funcion

        # Corrientes por circuito (manteniendo tu lógica)
        current_1WTG = (((wtg_unit_power / 1000) / 0.95) / (math.sqrt(3) * medium_voltage_level / 1000))
        current_2WTG = (((2 * wtg_unit_power / 1000) / 0.95) / (math.sqrt(3) * medium_voltage_level / 1000))
        current_3WTG = (((3 * wtg_unit_power / 1000) / 0.95) / (math.sqrt(3) * medium_voltage_level / 1000))

        # Resistencias (Ohm/km)
        resist_120mm2 = 0.325
        resist_300mm2 = 0.130
        resist_630mm2 = 0.063

        # Listas para devolver valores por SET
        perdidas_3WTG_por_set = []
        perdidas_2WTG_por_set = []
        perdidas_1WTG_por_set = []
        supply_120mm2_3WTGs = []
        supply_300mm2_3WTGs = []
        supply_630mm2_3WTGs = []
        supply_120mm2_2WTGs = []
        supply_300mm2_2WTGs = []
        supply_630mm2_1WTGs = []

        for r in resumen_sets:
            dist_media = r["Dist_Media_WTGs"]
            dist_min = r["Dist_Min"]
            dist_max = r["Dist_Max"]

            if dist_media is None or dist_min is None or dist_max is None:
                # Si el SET tiene 0-1 WTGs, mete ceros
                perdidas_3WTG_por_set.append(0.0)
                perdidas_2WTG_por_set.append(0.0)
                perdidas_1WTG_por_set.append(0.0)
                supply_120mm2_3WTGs.append(0.0)
                supply_300mm2_3WTGs.append(0.0)
                supply_630mm2_3WTGs.append(0.0)
                supply_120mm2_2WTGs.append(0.0)
                supply_300mm2_2WTGs.append(0.0)
                supply_630mm2_1WTGs.append(0.0)
                continue

            # Aplica factor de seguridad
            d_btw = dist_media * safety_factor_length
            d_mean = ((dist_max + dist_min) / 2) * safety_factor_length

            # Circuitos de 3 WTGs (manteniendo tu lógica)
            p_1_3 = 3 * (current_1WTG ** 2) * (d_btw * resist_120mm2)
            s_120_3 = d_btw * 3
            p_2_3 = 3 * (current_2WTG ** 2) * (d_btw * resist_300mm2)
            s_300_3 = d_btw * 3
            p_3_3 = 3 * (current_3WTG ** 2) * (d_mean * resist_630mm2)
            s_630_3 = d_mean * 3
            perdidas_3WTG_por_set.append(p_1_3 + p_2_3 + p_3_3)
            supply_120mm2_3WTGs.append(s_120_3)
            supply_300mm2_3WTGs.append(s_300_3)
            supply_630mm2_3WTGs.append(s_630_3)

            # Circuitos de 2 WTGs
            p_1_2 = 3 * (current_1WTG ** 2) * (d_btw * resist_120mm2)
            s_120_2 = d_btw * 3
            p_2_2 = 3 * (current_2WTG ** 2) * (d_mean * resist_630mm2)
            s_300_2 = d_mean * 3  # Mantengo tu lógica actual
            perdidas_2WTG_por_set.append(p_1_2 + p_2_2)
            supply_120mm2_2WTGs.append(s_120_2)
            supply_300mm2_2WTGs.append(s_300_2)

            # Circuitos de 1 WTG
            p_1_1 = 3 * (current_1WTG ** 2) * (d_mean * resist_630mm2)
            s_630_1 = d_mean * 3
            perdidas_1WTG_por_set.append(p_1_1)
            supply_630mm2_1WTGs.append(s_630_1)

        return (
            perdidas_3WTG_por_set,
            perdidas_2WTG_por_set,
            perdidas_1WTG_por_set,
            supply_120mm2_3WTGs,
            supply_300mm2_3WTGs,
            supply_630mm2_3WTGs,
            supply_120mm2_2WTGs,
            supply_300mm2_2WTGs,
            supply_630mm2_1WTGs
        )


    ################################################MENÚ OPCIONES###########################################
    opcion = 0
    def menu_opciones():
        print('Opción 1: Encuentra el óptimo con criterio de distancia')
        print('Opción 2: Coordenadas de SETs dadas.')
        print('Opción 3: Encuentra el óptimo de SETs modudulares de ', max_trafo_power/10**6,' MVA')
        print('Opción 4: Coordenadas de SETs dadas. SETs modudulares de ', max_trafo_power/10**6,' MVA.')
        print('Opción 5: Todas las opciones + Comparativa')
        global opcion
        opcion=int(input('Opción elegida: '))
        return opcion
    ################################################FUNCIONES###########################################
    ##Barrido del 1 a n_clusters_max (Para la opción 1)
    def barrido_nsets(n_clusters_max,opcion,descripcion_opcion):
        criterio_distancia=1
        sets_coord = calculo_centros(wtg_info, n_clusters_max)
        resumen_sets,opcion,n_clusters_max, wtg_fuera_radio,wtgs_por_set_filtrada=calculos_comunes_criterio_distancia(opcion,n_clusters_max,descripcion_opcion,criterio_distancia, wtg_info, sets_coord, distancia_radio_desde_SET)

        return criterio_distancia, descripcion_opcion, n_clusters_max,wtg_fuera_radio,sets_coord

    ###############################CALCULOS COMUNES A TODAS LAS OPCIONES#############################
    #Una vez se saben la distribución de WTGs a cada SET, los cálculos son iguales
    def calculos_comunes(sets_coord, criterio_distancia,wtgs_por_set_filtrada,opcion,n_clusters_max,wtg_fuera_radio):

        wtg_en_sets, asignaciones, resultados = asignar_wtgs_a_sets(sets_coord, wtg_info, usar_indice_desde_1=True)
        resumen_sets_funcion = resumen_por_set_completo(sets_coord, wtg_info, wtg_en_sets, usar_vecino_mas_cercano=True)
        # print('resumen_sets_funcion',resumen_sets_funcion)
        # print('wtg_en_sets',wtg_en_sets)


    #  --- NUEVO: alinear longitudes antes de calcular columnas/DF ---
        # Si hay discrepancia, recorta/filtra sets_coord con la máscara de "sets no vacíos"
        if len(sets_coord) != len(wtgs_por_set_filtrada):
            mask_no_vacio = [len(sublista) > 0 for sublista in wtgs_por_set_filtrada]
            # Si wtgs_por_set_filtrada ya está filtrada (todas no vacías), usa su longitud como referencia
            if all(mask_no_vacio):
                # Usa solo los primeros N centroides si sobran (o re-ordena si procede)
                sets_coord = np.array(sets_coord, dtype=float)[:len(wtgs_por_set_filtrada)]
            else:
                # Aplica la misma máscara sobre sets_coord
                sets_coord = np.array(sets_coord, dtype=float)[mask_no_vacio]
                # Y filtra también la lista para que coincidan 1 a 1
                wtgs_por_set_filtrada = [s for s in wtgs_por_set_filtrada if len(s) > 0]
        # ----------------------------------------------------------------


        n_sets=len(sets_coord)
        # print('Number of SETs',n_sets)

        # Contar elementos en cada SET
        conteo_por_set = [len(lista_wtgs) for lista_wtgs in wtgs_por_set_filtrada]

        # Mostrar resultados
        lista_cantidad=[]
        for i, cantidad in enumerate(conteo_por_set):
            # print(f"SET {i} tiene {cantidad} WTGs")
            lista_cantidad.append(cantidad)

        # print('Cantidad WTGs per SET', lista_cantidad)
        pot_per_SET_MW = [valor * wtg_unit_power/(10**6) for valor in lista_cantidad]
        # print('Potencia en MW per SET', pot_per_SET_MW)

        ##Cálculo cantidad de barras de media tensión
        number_busbar=[(((valor*10**6)/(math.sqrt(3)*medium_voltage_level))/busbar_limitation) for valor in pot_per_SET_MW]
        number_busbar_redondeados_arriba = [math.ceil(valor) for valor in number_busbar]
        n_totales_busbar=sum(number_busbar_redondeados_arriba)
        # print('Número busbar necesarios en cada SET (redondeados hacia arriba)', number_busbar_redondeados_arriba)
        # print('Número totales busbar', n_totales_busbar)
        ##Cálculo cantidad de transformadores
        number_transformers=[((valor*10**6)/max_trafo_power) for valor in pot_per_SET_MW]
        number_transformers_redondeados_arriba = [math.ceil(valor) for valor in number_transformers]
        n_totales_transformers=sum(number_transformers_redondeados_arriba)
        # print('Número transformadores necesarios en cada SET (redondeados hacia arriba)', number_transformers_redondeados_arriba)
        # print('Número totales transfomadores', n_totales_transformers)

        ##Cálculo CAPEX
        if criterio_distancia==1:
            CAPEX = cost_MV_Collector_system_10_5_km*n_wtg*wtg_unit_power+n_totales_busbar*cost_busbar_2500_A+n_totales_transformers*cost_transformer_300_MVA+n_sets*cost_SET
        else:
            CAPEX = cost_MV_Collector_system_300_MVA *n_wtg*wtg_unit_power+n_totales_busbar*cost_busbar_2500_A+n_totales_transformers*cost_transformer_300_MVA+n_sets*cost_SET
            
        #GENERACIÓN DE EXCEL
        # Convertir cada coordenada a string para la columna
        sets_coord_str = [f"[{x[0]:.6f}, {x[1]:.6f}]" for x in sets_coord]
        capex_column = [""] * len(sets_coord_str)  # O usa la longitud de la columna que defina el número de filas

        # Formatear CAPEX con comas para miles y punto para decimales
        capex_formateado = f"{CAPEX:,.2f}"  # Esto ya usa el formato inglés
        coord_x = [coord[0] for coord in sets_coord] # Coordenadas X de las SETs
        coord_y = [coord[1] for coord in sets_coord] # Coordenadas Y de las SETs

        if len(capex_column) >= 1:
            capex_column[0] = capex_formateado

            df = pd.DataFrame({
            "Coord X SET": coord_x,
            "Coord Y SET": coord_y,
            "Cantidad por SET": lista_cantidad,
            "Potencia por SET (MW)": pot_per_SET_MW,
            "Busbars por SET": number_busbar_redondeados_arriba,
            "Transformadores por SET": number_transformers_redondeados_arriba,
            "CAPEX": capex_column
        })

        if opcion==5:
                pass

    ###############################################################################

        elif opcion!=1 and wtg_fuera_radio==0:
            excel_filename = "RESULTADOS/" +str(descripcion_opcion)+  ", opción " +str(opcion)+".xlsx" # Se guardan los excels en la carpeta 'RESULTADOS'
            df.to_excel(excel_filename, index=False)
            print(f"Archivo Excel generado: {excel_filename}")

        elif opcion!=1 and wtg_fuera_radio==1:
            excel_filename = "RESULTADOS/WTGs fuera de radio/" +str(descripcion_opcion)+  ", opción " +str(opcion)+".xlsx" # Se guardan los excels en la carpeta 'RESULTADOS'
            df.to_excel(excel_filename, index=False)
            print(f"Archivo Excel generado: {excel_filename}")
        

        elif opcion==1 and wtg_fuera_radio==0:
            excel_filename = "RESULTADOS/" +str(descripcion_opcion)+ " "+ str(n_clusters_max) + " SETs, opción " +str(opcion)+".xlsx" # Se guardan los excels en la carpeta 'RESULTADOS'
            df.to_excel(excel_filename, index=False)
            print(f"Archivo Excel generado: {excel_filename}")

        elif opcion==1 and wtg_fuera_radio==1:
            excel_filename = "RESULTADOS//WTGs fuera de radio/" +str(descripcion_opcion)+ " "+ str(n_clusters_max) + " SETs, opción " +str(opcion)+".xlsx" # Se guardan los excels en la carpeta 'RESULTADOS'
            df.to_excel(excel_filename, index=False)
            print(f"Archivo Excel generado: {excel_filename}")

        # Retornar resultados si los quieres usar fuera
        return {
            "wtgs_por_set": wtgs_por_set_filtrada,
            "cantidad_por_set": lista_cantidad,
            "potencia_por_set_MW": pot_per_SET_MW,
            "busbars_por_set": number_busbar_redondeados_arriba,
            "transformadores_por_set": number_transformers_redondeados_arriba,
            "resumen_sets":resumen_sets_funcion

        }

    #######CÁLCULOS CENTRO
    def calculo_centros(wtg_info, n_clusters):
        coordenadas=np.array([[x[1], x[2]] for x in wtg_info])

        #Aplicar KMeans para obetener tantos centroides como clusters deseados
        # KMeans es un algoritmo de clustering que agrupa datos en K clusters según su similitud.
        # Asigna puntos al cluster más cercano, recalcula los centros, y repite hasta converger.
        kmeans = KMeans(n_clusters, random_state=42)
        kmeans.fit(coordenadas)
        centroides = kmeans.cluster_centers_
        labels = kmeans.labels_
        return centroides

    #######CALCULOS COMUNES CRITERIO DE DISTANCIA#########
    def calculos_comunes_criterio_distancia(opcion,n_clusters_max,descripcion_opcion,criterio_distancia, wtg_info, sets_coord, distancia_radio_desde_SET):
        wtg_en_sets = defaultdict(list)
        for idx, set_coord in enumerate(sets_coord):
            for fila in wtg_info:
                id_wtg, x, y = fila[0], fila[1], fila[2]
                distancia = math.sqrt((x - set_coord[0])**2 + (y - set_coord[1])**2)
            
                if 0 < distancia < distancia_radio_desde_SET:
                    wtg_en_sets[id_wtg].append(idx + 1)

        # Mostrar WTGs que están en más de un SET
        if len(sets_coord) > 2:
            wtgs_duplicados = {wtg: sets for wtg, sets in wtg_en_sets.items() if len(sets) > 1}
            if wtgs_duplicados:
                for wtg, sets in wtgs_duplicados.items():
                    a=1
            else:
                a=0

        #Calcular de cual de las SETs está más cerca
        wtg_mas_cercano = []

        wtgs_fuera_de_todas = []
        for wtg in wtg_info:
            id_wtg, x, y = wtg[0], wtg[1], wtg[2]
            distancia_min = float('inf')
            set_mas_cercano = None

            for idx, set_coord in enumerate(sets_coord):
                distancia = math.sqrt((x - set_coord[0])**2 + (y - set_coord[1])**2)
                if distancia < distancia_min:
                    distancia_min = distancia
                    set_mas_cercano = idx  # o puedes guardar el centroide directamente
                    
                    if distancia>distancia_radio_desde_SET:
                        wtg_fuera_radio=1

                    elif distancia<distancia_radio_desde_SET:
                        wtg_fuera_radio=0
                

            wtg_mas_cercano.append((id_wtg, set_mas_cercano, distancia_min))    
            # Si la distancia mínima es mayor que el radio, añadir a la lista
            if distancia_min > distancia_radio_desde_SET:           
                coord_set_cercano = (float(sets_coord[set_mas_cercano][0]), float(sets_coord[set_mas_cercano][1]))
                wtgs_fuera_de_todas.append((id_wtg, coord_set_cercano))

        resumen_sets=0
        if len(wtgs_fuera_de_todas)!=0:
            wtgs_por_set_cercano = defaultdict(list)

            for wtg_id, coord_set in wtgs_fuera_de_todas:
                wtgs_por_set_cercano[coord_set].append(wtg_id)

            # Crear nombres para las SETs si no los tienes
            sets_nombres = [f"SET_{i+1}" for i in range(len(sets_coord))]

            # Agrupar WTGs fuera del radio por SET más cercana
            wtgs_por_set_cercano = defaultdict(list)

            for wtg_id, coord_set in wtgs_fuera_de_todas:
                # Buscar el índice de la coordenada en sets_coord
                idx = None
                for i, sc in enumerate(sets_coord):
                    if float(sc[0]) == coord_set[0] and float(sc[1]) == coord_set[1]:
                        idx = i
                        break
                if idx is not None:
                    wtgs_por_set_cercano[sets_nombres[idx]].append(wtg_id)

            # Convertir a lista organizada (opcional)
            lista_organizada = [(nombre_set, wtgs) for nombre_set, wtgs in wtgs_por_set_cercano.items()]
            
            # Crear lista con SET, coordenada y número de WTGs
            resumen_sets = [(nombre_set, len(lista_wtgs)) for nombre_set, lista_wtgs in wtgs_por_set_cercano.items()]
            # Mostrar la lista
            print("\nResumen por SET más cercana:")
            print(resumen_sets)

        # Crear lista de listas: cada índice representa un SET
        wtgs_por_set = [[] for _ in range(len(sets_coord))]
        for wtg_id, set_idx, dist in wtg_mas_cercano:
            wtgs_por_set[set_idx].append(wtg_id) 
        # wtgs_por_set_filtrada = [sublista for sublista in wtgs_por_set if sublista] #Elimina sublistas vacías (cuando solo hay 1 SET)

        # 1) Máscara de SETs no vacíos
        mask_no_vacio = [len(sublista) > 0 for sublista in wtgs_por_set]

        # 2) Filtrar sets_coord con la misma máscara (para que todas las columnas tengan la misma longitud)
        sets_coord_filtrada = np.array(sets_coord)[mask_no_vacio]  # si sets_coord ya es np.array, basta con sets_coord[mask_no_vacio]

        # 3) Lista de WTGs por SET sin vacíos (como ya hacías)
        wtgs_por_set_filtrada = [sublista for sublista in wtgs_por_set if sublista]




        

        if criterio_distancia==1: #Ejecuta esto solo con las opciones 1 y 2
            conteo_por_set = [len(lista_wtgs) for lista_wtgs in wtgs_por_set_filtrada]

            # Mostrar resultados
            lista_cantidad=[]
            for i, cantidad in enumerate(conteo_por_set):
                lista_cantidad.append(cantidad)

            # Visualización
            fig, ax = plt.subplots(figsize=(10, 8))#ESTO NO SÉ SI VA AQUÍ
            # Crear diccionario para acceder rápidamente a coordenadas por nombre
            wtg_coord_dict = {wtg[0]: (wtg[1], wtg[2]) for wtg in wtg_info}
            # Colores para los grupos
            colors = plt.cm.get_cmap('tab10', len(wtgs_por_set_filtrada))

            # Lista para guardar los handles de los grupos de WTGs
            scatter_wtg_handles = []

            # Dibujar cada grupo
            for i, grupo in enumerate(wtgs_por_set_filtrada):
                wtg_xs, wtg_ys = [], []

                for wtg in grupo:
                    if wtg in wtg_coord_dict:
                        x, y = wtg_coord_dict[wtg]
                        wtg_xs.append(x)
                        wtg_ys.append(y)
                scatter_wtg=ax.scatter(wtg_xs, wtg_ys, label=f'SET {i+1} ({len(grupo)} WTGs)', s=30, color=colors(i))
                scatter_wtg_handles.append(scatter_wtg)  # Guardar cada scatter para que salgan en la leyenda cada SET con el número de WTGs que tiene
                
            # Dibujar puntos de SETs
            # scatter_sets=ax.scatter(sets_coord[:, 0], sets_coord[:, 1], color='black', marker='X', s=100, label='SETs')
            scatter_sets=ax.scatter(sets_coord_filtrada[:, 0], sets_coord_filtrada[:, 1], color='black', marker='X', s=100, label='SETs')

        
        ax.legend(handles=scatter_wtg_handles + [scatter_sets])
        ax.set_xlabel('Coordenada X')
        ax.set_ylabel('Coordenada Y')
        ax.set_title("Opción" + str(opcion) +". " + str(descripcion_opcion) +  ".")
        ax.grid(True)
        plt.axis('equal')
        plt.tight_layout()

        if wtg_fuera_radio == 1:# Si hay WTGs fuera del radio, dibujar el recuadro Y EL CÍRCULO 
            data = []
            for set_name, items in resumen_sets:
                row = [set_name] + (items if isinstance(items, list) else [items])
                data.append(row)
            # Creamos encabezados dinámicos
            max_cols = max(len(row) for row in data)
            headers = ['SET'] + [f'Item{i}' for i in range(1, max_cols)]

            df = pd.DataFrame(data, columns=headers)


            # Se divide entre 2 para ver cuantos circuitos de 2WTGs habrá, si el numero es impar tendrá que haber sircuitos de 1WTG
            divisor = 2

            # Añadimos columnas calculadas
            df['Nºcircuitos de 2 WTGs'] = df['Item1'].astype(int) // divisor
            df['Nºcircuitos de 1 WTG'] = df['Item1'].astype(int) % divisor

            if opcion==5:
                pass

            elif opcion!=1:
                excel_filename = "RESULTADOS/WTGs fuera de radio/SETs/" +str(descripcion_opcion)+  ", opción " +str(opcion)+"_WTGs per SETS.xlsx" # Se guardan los excels en la carpeta 'RESULTADOS'
                df.to_excel(excel_filename, index=False)
                print(f"Archivo Excel generado: {excel_filename}")
            
            else:
                excel_filename = "RESULTADOS/WTGs fuera de radio/SETs/" +str(descripcion_opcion)+ " "+ str(n_clusters_max) + "_WTGs per SETS, opción " +str(opcion)+"PRUEBA SETS.xlsx" # Se guardan los excels en la carpeta 'RESULTADOS'
                df.to_excel(excel_filename, index=False)
                print(f"Archivo Excel generado: {excel_filename}")

            # Dibuja un círculo de radio 'distancia_radio_desde_SET' centrado en cada SET
            for x, y in sets_coord:
                circle = Circle((x, y), distancia_radio_desde_SET, color='blue', fill=False, linestyle='--', linewidth=2, alpha=0.7)
                ax.add_patch(circle)

            # Añadir una entrada a la leyenda para representar los círculos
                legend_circle = Patch(edgecolor='blue', facecolor='none', linestyle='--', linewidth=2,
                                label=f'Radius {distancia_radio_desde_SET/1000} km')


                # Añadir leyenda con todos los elementos
                ax.legend(handles=scatter_wtg_handles + [scatter_sets, legend_circle])
                ax.set_xlabel('Coordenada X')
                ax.set_ylabel('Coordenada Y')

                xlim = ax.get_xlim()
                ylim = ax.get_ylim()
                ancho = (xlim[1] - xlim[0]) * 0.25 #Ancho del recuadro
                alto = (ylim[1] - ylim[0]) * 0.15 #Alto del recuadro
                x_rect = xlim[1] - ancho - (xlim[1] - xlim[0]) * 0.002 #Coloca la coordenada X del recuadro
                y_rect = ylim[0] + (ylim[1] - ylim[0]) * 0.01 #Coloca la coordenada Y del recuadro   # abajo

                rect = patches.FancyBboxPatch(
                    (x_rect, y_rect), ancho, alto,
                    boxstyle="round,pad=0.3",
                    linewidth=2, edgecolor='red',
                    facecolor='lightyellow', alpha=0.9
                )
                ax.add_patch(rect)
                ax.text(
                    x_rect + ancho / 2,
                    y_rect + alto / 2,
                    "CAUTION!\n WTGs out of range\nOutside the circle means\ncircuits with 1 or 2 WTGs",#Texto del recuadro
                    ha='center', va='center', fontsize=11, color='black'
                )

        if opcion == 2:
            plt.savefig("IMAGENES/" +str(descripcion_opcion)+  ", opción " +str(opcion)+".png", dpi=300)# Guarda la imagen #GUARDA LA IMAGEN EN LA CARPETA IMAGENES
        if opcion == 1: 
            plt.savefig("IMAGENES/" +str(descripcion_opcion)+" "+ str(n_clusters_max) + " SETs,  opción " +str(opcion)+" .png", dpi=300)# Guarda la imagen #GUARDA LA IMAGEN EN LA CARPETA IMAGENES                                          

        # resumen_sets_funcion=calculos_comunes(sets_coord, criterio_distancia,wtgs_por_set_filtrada,opcion,n_clusters_max,wtg_fuera_radio)
        resumen_sets_funcion=calculos_comunes(sets_coord_filtrada, criterio_distancia,wtgs_por_set_filtrada,opcion,n_clusters_max,wtg_fuera_radio)

        return resumen_sets_funcion,opcion,n_clusters_max, wtg_fuera_radio,wtgs_por_set_filtrada



    #################################################################################################################################
    ##########OPCIÓN 1 se establece los puntos centrales de los WTGs##################

    def opcion1(descripcion_opcion, opcion, n_clusters_max):
        n_clusters_op1A=n_clusters_max
        sets_coord = calculo_centros(wtg_info, n_clusters_op1A)
        print('sets_coord EN LA FUNCION DE OPCION 1',sets_coord)

        criterio_distancia=1
        resumen_sets,opcion,n_clusters_max, wtg_fuera_radio,wtgs_por_set_filtrada=calculos_comunes_criterio_distancia(opcion,n_clusters_max,descripcion_opcion,criterio_distancia, wtg_info, sets_coord, distancia_radio_desde_SET)
        return criterio_distancia, descripcion_opcion, opcion,resumen_sets,wtg_fuera_radio,wtgs_por_set_filtrada

    ##########OPCIÓN 2 te dan como input las coordenadas de las SETs##########################################################################################
    def opcion2(descripcion_opcion, opcion):
        n_clusters_max=0
        criterio_distancia=1
        #Lee el archivo donde se encuentran las coordenadas de las SETs dadas. 
        # El formato debe de ser: 1ºcolumna: Coordenada X, 2ºcolumna: Coordenada Y 
        # wb=load_workbook("SET_Coord_prueba_22092025.xlsx", read_only=True)#El excell debe estar en la misma carpeta que este código y CERRADO
        wb=load_workbook("SET_Coord_test_run.xlsx", read_only=True)
        sheet=wb.active

        #Leer todas las filas del excel y almacenarlas como listas dentro de una lista global llamada wtg_info
        sets_info=[]
        for fila in sheet.iter_rows(values_only=True):
            sets_info.append(list(fila))
        print(sets_info)

        #Calcula el número de WTGs que hay
        n_set=len(sets_info)
        sets_coord=np.array(sets_info)
        resumen_sets,opcion,n_clusters_max, wtg_fuera_radio,wtgs_por_set_filtrada=calculos_comunes_criterio_distancia(opcion,n_clusters_max,descripcion_opcion,criterio_distancia, wtg_info, sets_coord, distancia_radio_desde_SET)

        return criterio_distancia, resumen_sets, descripcion_opcion, opcion,wtgs_por_set_filtrada

    ############OPCIÓN 3########### AGRUPA EL PARQUE EN SUBESTACIONES DE MÁXIMO 300MVA##########
    #NOTA: SE QUEDAN UNOS 3 WTGS MEZCLADOS CON LOS DE AL LADO, PERO POR AHORA NO AFECTA A NADA PORQUE NO SE CALCULA LA MEDIA TENSIÓN :)
    def opcion3(descripcion_opcion, opcion):
        criterio_distancia=0
        n_clusters_max=0
        n_clusters_op1B = math.ceil(n_wtg*wtg_unit_power/max_trafo_power)#Cálcula y redondea hacia arriba

        # Parámetros
        max_WTG_per_SET = math.floor(max_trafo_power/wtg_unit_power)#Cálcula el nº de WTGs máximo por SET y redondea hacia abajo
        num_clusters = math.ceil(n_wtg*wtg_unit_power/max_trafo_power)#Cálcula el número de clusters (SETs) y redondea hacia arriba

        # Leer el archivo Excel. #El formato debe de ser: 1ºcolumna: Nombre WTG, 2ºcolumna: Coordenada X, 3columna: Coordenada Y
        
        df = pd.read_excel(coordenadas_wtg_excel, engine="openpyxl", header=None)
        wtg_info = df.values.tolist()

        # Limpiar y convertir coordenadas
        wtg_info_clean = [[entry[0], float(entry[1]), float(entry[2])] for entry in wtg_info]
        coords = np.array([[entry[1], entry[2]] for entry in wtg_info_clean])

        # Inicializar centroides con KMeans para obtener puntos de partida
        kmeans = KMeans(n_clusters=num_clusters, random_state=0)
        kmeans.fit(coords)
        initial_centroids = kmeans.cluster_centers_

        # Reorganización respetando el límite de WTGs por grupo
        assigned_groups = [[] for _ in range(num_clusters)]
        remaining_indices = list(range(len(coords)))

        # Asignar WTGs al centroide más cercano que tenga espacio disponible
        while remaining_indices:
            distances = np.linalg.norm(coords[remaining_indices][:, np.newaxis] - initial_centroids, axis=2)
            new_remaining_indices = []
            for i, idx in enumerate(remaining_indices):
                sorted_centroids = np.argsort(distances[i])
                assigned = False
                for c in sorted_centroids:
                    if len(assigned_groups[c]) < max_WTG_per_SET:
                        assigned_groups[c].append(wtg_info_clean[idx])
                        assigned = True
                        break
                if not assigned:
                    new_remaining_indices.append(idx)
            if len(new_remaining_indices) == len(remaining_indices):
                print("No se pudo asignar algunos WTGs debido a la falta de espacio en todos los grupos.")
                break
            remaining_indices = new_remaining_indices

        # Calcular nuevos centroides
        new_centroids = []
        for grupo in assigned_groups:
            xs = [p[1] for p in grupo]
            ys = [p[2] for p in grupo]
            new_centroids.append((np.mean(xs), np.mean(ys)))
        new_centroids = np.array(new_centroids)

        # Visualización
        fig, ax = plt.subplots(figsize=(10, 8))
        for i, grupo in enumerate(assigned_groups):
            xs = [p[1] for p in grupo]
            ys = [p[2] for p in grupo]
            ax.scatter(xs, ys, label=f'SET {i+1} ({len(grupo)} WTGs)', s=30)
            # for punto in grupo:
            #     plt.text(punto[1]+10, punto[2]+10, punto[0], fontsize=6)#Esto pone el nombre de cada WTG en el dibujo

    ####################CAMBIAR ESTO PARA INDICAR CUAL EL LÍMITE DE POTENICA EN ESTE CASO#########################################################################

    # Añadir una entrada a la leyenda para representar los círculos

        # Añadir leyenda con todos los elementos

        xlim = ax.get_xlim()
        ylim = ax.get_ylim()
        ancho = (xlim[1] - xlim[0]) * 0.25 #Ancho del recuadro
        alto = (ylim[1] - ylim[0]) * 0.15 #Alto del recuadro
        x_rect = xlim[1] - ancho - (xlim[1] - xlim[0]) * 0.002 #Coloca la coordenada X del recuadro
        y_rect = ylim[0] + (ylim[1] - ylim[0]) * 0.01 #Coloca la coordenada Y del recuadro   # abajo

        rect = patches.FancyBboxPatch(
            (x_rect, y_rect), ancho, alto,
            boxstyle="round,pad=0.3",
            linewidth=2, edgecolor='red',
            facecolor='lightyellow', alpha=0.9
        )

        ax.add_patch(rect)
        ax.text(
            x_rect + ancho / 2,
            y_rect + alto / 2,
            "CAUTION!\n The limit for limit\n current "+str(busbar_limitation)+" is\n" +str(max_WTG_per_SET)+ " WTGs",#Texto del recuadro
            ha='center', va='center', fontsize=11, color='black'
        )
    #############################################################################################
        ax.scatter(new_centroids[:, 0], new_centroids[:, 1], color='black', marker='X', s=100, label='SETs')
        ax.set_title("Opción" + str(opcion) +". " + str(descripcion_opcion) +  ".")
        ax.grid(True)
        ax.legend()
        ax.set_xlabel('Coordenada X')
        ax.set_ylabel('Coordenada Y')
        plt.tight_layout()

    #####
        plt.savefig("IMAGENES/" +str(descripcion_opcion)+  ", opción " +str(opcion)+".png", dpi=300)# Guarda la imagen #GUARDA LA IMAGEN EN LA CARPETA IMAGENES
        # plt.show()

        # Mostrar conteo por grupo
        group_sizes = [len(grupo) for grupo in assigned_groups]
        print("Número de WTGs por grupo:", group_sizes)
        sets_coord = new_centroids

        # Crear lista de listas: cada índice representa un SET
        wtgs_por_set = [[] for _ in range(len(sets_coord))]
        wtgs_por_set = assigned_groups
        # wtgs_por_set_filtrada = [sublista for sublista in wtgs_por_set if sublista] #Elimina sublistas vacías (cuando solo hay 1 SET)
        # calculos_comunes(sets_coord, criterio_distancia,wtgs_por_set_filtrada,opcion,n_clusters_max,wtg_fuera_radio)


        # 1) Máscara de SETs no vacíos
        mask_no_vacio = [len(sublista) > 0 for sublista in wtgs_por_set]

        # 2) Filtrar sets_coord con la misma máscara (para que todas las columnas tengan la misma longitud)
        sets_coord_filtrada = np.array(sets_coord)[mask_no_vacio]  # si sets_coord ya es np.array, basta con sets_coord[mask_no_vacio]

        # 3) Lista de WTGs por SET sin vacíos (como ya hacías)
        wtgs_por_set_filtrada = [sublista for sublista in wtgs_por_set if sublista]

        # 4) Pasar la versión filtrada a calculos_comunes(...)
        resumen_sets_funcion = calculos_comunes(
            sets_coord_filtrada,          # <-- ¡filtrado!
            criterio_distancia,
            wtgs_por_set_filtrada,
            opcion,
            n_clusters_max,
            wtg_fuera_radio
        )



        return descripcion_opcion, opcion,wtgs_por_set_filtrada

    ############OPCIÓN 4########### AGRUPA EL PARQUE EN SUBESTACIONES DE MÁXIMO 300MVA A UNAS COORDENADAS DADAS (no ha pasado nunca) ##########
    def opcion4(descripcion_opcion, opcion):
        max_wtg_per_set = math.floor(max_trafo_power/wtg_unit_power)#Redondea hacia abajo
        criterio_distancia=0
        n_clusters_max=0
        #Lee el archivo donde se encuentran las coordenadas de las SETs dadas
        #El formato debe de ser: 1ºcolumna: Coordenada X, 2ºcolumna: Coordenada Y 
        # wb=load_workbook("SET_Coord_prueba_300MVA.xlsx", read_only=True)#El excell debe estar en la misma carpeta que este código y CERRADO
        wb=load_workbook("SET_Coord_300MVA_test_run.xlsx", read_only=True)
        sheet=wb.active

        #Leer todas las filas dle excel y almacenarlas como listas dentro de una lista global llamada wtg_info
        sets_info=[]
        for fila in sheet.iter_rows(values_only=True):
            sets_info.append(list(fila))

        #Calcula el número de WTGs que hay
        n_set=len(sets_info)
        print('El número de SETs es: ', n_set)
        sets_coord=np.array(sets_info)
        
        # Leer coordenadas de WTGs. #El formato debe de ser: 1ºcolumna: Nombre WTG, 2ºcolumna: Coordenada X, 3ºcolumna: Coordenada Y
        wtg_df = pd.read_excel(coordenadas_wtg_excel, engine="openpyxl", header=None)
        wtg_info = wtg_df.values.tolist()
        wtg_info_clean = [[entry[0], float(entry[1]), float(entry[2])] for entry in wtg_info]
        wtg_coords = np.array([[entry[1], entry[2]] for entry in wtg_info_clean])

        # Leer coordenadas de SETs. #El formato debe de ser: 1ºcolumna: Coordenada X, 2ºcolumna: Coordenada Y 
        # wb = load_workbook("SET_Coord_prueba_300MVA.xlsx", read_only=True)
        wb = load_workbook("SET_Coord_300MVA_test_run.xlsx", read_only=True)
        sheet = wb.active
        sets_info = [list(row) for row in sheet.iter_rows(values_only=True)]
        sets_coords = np.array(sets_info)

        # Inicializar grupos
        wtgs_por_set = [[] for _ in range(len(sets_coords))]
        set_capacities = [0] * len(sets_coords)
        remaining_indices = list(range(len(wtg_coords)))

        # Asignar WTGs al SET más cercano que tenga espacio disponible
        while remaining_indices:
            distances = np.linalg.norm(wtg_coords[remaining_indices][:, np.newaxis] - sets_coords, axis=2)
            new_remaining = []
            for i, idx in enumerate(remaining_indices):
                sorted_sets = np.argsort(distances[i])
                assigned = False
                for s in sorted_sets:
                    if set_capacities[s] < max_wtg_per_set:
                        wtgs_por_set[s].append(wtg_info_clean[idx])
                        set_capacities[s] += 1
                        assigned = True
                        break
                if not assigned:
                    new_remaining.append(idx)
            if len(new_remaining) == len(remaining_indices):
                print("No se pudo asignar algunos WTGs debido a falta de capacidad.")
                break
            remaining_indices = new_remaining

        # Mostrar distribución final
        # group_sizes = [len(grupo) for grupo in wtgs_por_set]
        # print("Distribución final de WTGs por SET:", group_sizes)

        
        # Visualización
        fig, ax = plt.subplots(figsize=(10, 8))
        colors = plt.cm.get_cmap('tab10', len(wtgs_por_set))
        for i, grupo in enumerate(wtgs_por_set):
            xs = [p[1] for p in grupo]
            ys = [p[2] for p in grupo]
            ax.scatter(xs, ys, label=f'SET {i+1} ({len(grupo)} WTGs)', s=30, color=colors(i))
            # for punto in grupo:
            #     plt.text(punto[1]+10, punto[2]+10, punto[0], fontsize=6)# Esto pone el nombre de cada WTG en el dibujo

        ax.scatter(sets_coords[:, 0], sets_coords[:, 1], color='black', marker='X', s=100, label='SETs')
        # ax.scatter(sets_coords_filtrada[:, 0], sets_coords_filtrada[:, 1], color='black', marker='X', s=100, label='SETs')
        ax.set_title("Opción" + str(opcion) +". " + str(descripcion_opcion) +  ".")
        ax.grid(True)
        ax.legend()
        ax.set_xlabel('Coordenada X')
        ax.set_ylabel('Coordenada Y')
        plt.tight_layout()

    #####NOTA: No creo que sea necesario diujar el circulo de distancia máxima porque aquí no aplica. Al haber más SETs todos los WTGs estarán cerca de alguna.
    # # Dibuja un círculo de radio 'distancia_radio_desde_SET' centrado en cada SET
    #     for x, y in sets_coords:
    #         circle = plt.Circle((x, y), distancia_radio_desde_SET, color='blue', fill=False, linestyle='--', linewidth=2, alpha=0.7)
    #         ax.add_patch(circle)

        plt.savefig("IMAGENES/" +str(descripcion_opcion)+  " ,opción " +str(opcion)+".png", dpi=300) #GUARDA LA IMAGEN EN LA CARPETA IMAGENES
        # plt.show()
        # Crear lista de listas: cada índice representa un SET
        wtgs_por_set_filtrada = [sublista for sublista in wtgs_por_set if sublista] #Elimina sublistas vacías (cuando solo hay 1 SET)
        # calculos_comunes(sets_coord, criterio_distancia,wtgs_por_set_filtrada,opcion,n_clusters_max,wtg_fuera_radio)



        # ✓ FILTRADO COHERENTE (una sola vez)
        mask_no_vacio = [len(grupo) > 0 for grupo in wtgs_por_set]
        sets_coords_filtrada = np.array(sets_coords, dtype=float)[mask_no_vacio]
        wtgs_por_set_filtrada = [grupo for grupo, keep in zip(wtgs_por_set, mask_no_vacio) if keep]

        # Plot usando siempre coords filtradas
        ax.scatter(sets_coords_filtrada[:, 0], sets_coords_filtrada[:, 1], color='black', marker='X', s=100, label='SETs')

        resumen_sets_funcion = calculos_comunes(
            sets_coords_filtrada,          # <-- ¡filtrado!
            criterio_distancia,
            wtgs_por_set_filtrada,
            opcion,
            n_clusters_max,
            wtg_fuera_radio
        )

        sets_coord_op4 = sets_coords_filtrada
        wtgs_por_set_filtrada_op4 = wtgs_por_set_filtrada


        # Si quieres devolver la lista filtrada:
        return criterio_distancia, descripcion_opcion, opcion, wtgs_por_set_filtrada


        # sets_coord_op4 = sets_coords_filtrada
        # wtgs_por_set_filtrada_op4 = wtgs_por_set_filtrada

        
        # # return criterio_distancia, descripcion_opcion, opcion
        # return criterio_distancia, descripcion_opcion, opcion, wtgs_por_set_filtrada_op4

    ############################
    wtg_opcion1 = 0
    wtg_opcion2 = 0
    wtg_opcion3 = 0
    wtg_opcion4 = 0

    descripcion_opcion=0

    opcion=0
    opcion=menu_opciones()
    n_clusters_max=5#Se elige con cuantas SETs se va a calcular el punto óptimo

    if opcion==1:
        wtg_fuera_radio=0

        descripcion_opcion = 'SET calculadas. Criterio 10.5km'

        # Elige si quieres barrer todos los SETs o un i concreto:
        #  - Para barrer todos los i: usa range(1, n_clusters_max + 1)
        #  - Para un i concreto: usa [i] (por ejemplo [3])
        for i in range(1, n_clusters_max + 1):
            # 1) Barrido para obtener n_clusters_op1A (igual que haces ahora)
            criterio_distancia, descripcion_opcion, n_clusters_max, wtg_fuera_radio, sets_coord = barrido_nsets(i, opcion, descripcion_opcion)
            n_clusters_op1A = n_clusters_max

            # 2) Ejecutar la opción 1 con ese número de SETs
            criterio_distancia, descripcion_opcion, opcion, resumen_sets_funcion_op1, wtg_fuera_radio, wtgs_por_set_filtrada_op1 = opcion1(descripcion_opcion, opcion, n_clusters_op1A)

            # 3) Construir el nombre base del Excel que generó calculos_comunes() para esta opción
            #    IMPORTANTE: este nombre debe coincidir EXACTAMENTE con el que calcula 'calculos_comunes()'
            base_excel_op1 = f"{descripcion_opcion} {i} SETs, opción {opcion}.xlsx"

            # 4) Elegir dónde buscar el Excel fuente (si hubo WTGs fuera de radio, el archivo está en esa subcarpeta)
            if wtg_fuera_radio == 1:
                archivos_a_procesar = [os.path.join("WTGs fuera de radio", base_excel_op1)]
            else:
                archivos_a_procesar = [base_excel_op1]

            # 5) Llamar a procesar_excels para generar: 
            #    - el Excel de resultados por circuito (completo)
            #    - el Excel de Pérdidas Totales (guardado en tu carpeta fija)
            resumen_op1 = procesar_excels(
                wtg_opcion1=wtg_fuera_radio,   # bandera usada para ajustar búsquedas/etiquetas, si la necesitas
                wtg_opcion2=0,
                wtg_opcion3=0,
                wtg_opcion4=0,
                carpeta_resultados='./RESULTADOS',
                resumen_sets_funcion=resumen_sets_funcion_op1,    # dict/list con "resumen_sets"
                archivo_final=os.path.join(r"./Resultados Circuitos",f"resultados_circuitos_opcion1_{i}_SETs.xlsx"),
                archivo_perdidas=os.path.join(r"./Pérdidas Totales",f"Pérdidas_Totales_opcion1_{i}_SETs.xlsx"),
                archivos_filtrar=archivos_a_procesar
            )

        
    elif opcion==2:
        wtg_fuera_radio=0
        descripcion_opcion = 'Coordenadas de SETs dadas. Criterio 10.5km'
        criterio_distancia, resumen_sets_funcion_op2, descripcion_opcion, opcion, wtgs_por_set_filtrada_op2 = opcion2(descripcion_opcion, opcion)

        base_excel_op2 = f"{descripcion_opcion}, opción {opcion}.xlsx"
        resumen_op2 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='./RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op2,
            archivo_final=os.path.join(r"./Software SETs/Resultados Circuitos",f"resumen_circuitos_opcion2.xlsx"),
            archivo_perdidas=os.path.join(r"./Software SETs/Pérdidas Totales",f"Pérdidas_Totales_opcion2.xlsx"),
            archivos_filtrar=[base_excel_op2]
        )


        
    elif opcion==3:
        wtg_fuera_radio=0
        descripcion_opcion = 'SET calculadas. Criterio límite de corriente en barras (SETs modulares pequeñas)'
        descripcion_opcion, opcion, wtgs_por_set_filtrada_op3 = opcion3(descripcion_opcion, opcion)
        sets_coord_op3 = np.array([[np.mean([p[1] for p in grupo]), np.mean([p[2] for p in grupo])] for grupo in wtgs_por_set_filtrada_op3])
        resumen_sets_funcion_op3 = calculos_comunes(sets_coord_op3, 0, wtgs_por_set_filtrada_op3, opcion, n_clusters_max, 0)
        base_excel_op3 = f"{descripcion_opcion}, opción {opcion}.xlsx"
        resumen_op3 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='./Software SETs/RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op3,
            archivo_final=os.path.join(r"./Software SETs/Resultados Circuitos",f"resumen_circuitos_opcion3.xlsx"),
            archivo_perdidas=os.path.join(r"./Software SETs/Pérdidas Totales",f"Pérdidas_Totales_opcion3.xlsx"),
            archivos_filtrar=[base_excel_op3]
        )
        
    elif opcion==4:
        wtg_fuera_radio=0
        descripcion_opcion='Coordenadas de SETs dadas. Criterio límite de corriente en barras (SETs modulares pequeñas)'
        criterio_distancia, descripcion_opcion, opcion, wtgs_por_set_filtrada_op4 = opcion4(descripcion_opcion, opcion)
        sets_coord_op4 = np.array([[np.mean([p[1] for p in grupo]), np.mean([p[2] for p in grupo])] for grupo in wtgs_por_set_filtrada_op4])
        resumen_sets_funcion_op4 = calculos_comunes(sets_coord_op4, 0, wtgs_por_set_filtrada_op4, opcion, n_clusters_max, 0)

        base_excel_op4 = f"{descripcion_opcion}, opción {opcion}.xlsx"
        resumen_op4 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='./Software SETs/RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op4,
            archivo_final=os.path.join(r"./Software SETs/Resultados Circuitos",f"resumen_circuitos_opcion4.xlsx"),
            archivo_perdidas=os.path.join(r"./Software SETs/Pérdidas Totales",f"Pérdidas_Totales_opcion4.xlsx"),
            archivos_filtrar=[base_excel_op4]
        )

    ####################################################################################################################################

    elif opcion == 5:

        # === RECOPILACIÓN FINAL ===
        resumen_global = []            # lista de dicts, se convertirá a DataFrame
        detalles_circuitos_dfs = []    # lista de DataFrames (detalle por SET)

        # --- Opción 1 (barrido 1..n_clusters_max) ---
        opcion = 1
        descripcion_opcion = 'SET calculadas. Criterio 10.5km'
        for i in range(1, n_clusters_max + 1):
            criterio_distancia, descripcion_opcion, opcion, resumen_sets_funcion_op1, wtg_fuera_radio, wtgs_por_set_filtrada_op1 = opcion1(descripcion_opcion, opcion, i)

            base_excel_op1 = f"{descripcion_opcion} {i} SETs, opción {opcion}.xlsx"
            if wtg_fuera_radio == 1:
                archivos_a_procesar = [os.path.join("WTGs fuera de radio", base_excel_op1)]
            else:
                archivos_a_procesar = [base_excel_op1]

            resumen_op1 = procesar_excels(
                wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
                carpeta_resultados='./Software SETs/RESULTADOS',
                resumen_sets_funcion=resumen_sets_funcion_op1,
                archivo_final=os.path.join(r"./Software SETs/Resultados Circuitos",f"resultados_circuitos_opcion1_{i}_SETs.xlsx"),
                archivo_perdidas=os.path.join(r"./Software SETs/Pérdidas Totales",f"Pérdidas_Totales_opcion1_{i}_SETs.xlsx"),
                archivos_filtrar=archivos_a_procesar
            )
            # Acumular en resumen_global
            for fila in resumen_op1["resumen_por_config"]:
                fila_rec = fila.copy()
                fila_rec["Opción"] = f"1 (SETs={i})"
                resumen_global.append(fila_rec)
            # Acumular detalle por SET
            detalles_circuitos_dfs.append(resumen_op1["detalle_por_set"])

        # --- Opción 2 ---
        opcion = 2
        descripcion_opcion = 'Coordenadas de SETs dadas. Criterio 10.5km'
        criterio_distancia, resumen_sets_funcion_op2, descripcion_opcion, opcion, wtgs_por_set_filtrada_op2 = opcion2(descripcion_opcion, opcion)

        base_excel_op2 = f"{descripcion_opcion}, opción {opcion}.xlsx"
        resumen_op2 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='./Software SETs/RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op2,
            archivo_final=os.path.join(r"./Software SETs/Resultados Circuitos",f"resumen_circuitos_opcion2.xlsx"),
            archivo_perdidas=os.path.join(r"./Software SETs/Pérdidas Totales",f"Pérdidas_Totales_opcion2.xlsx"),
            archivos_filtrar=[base_excel_op2]
        )
        for fila in resumen_op2["resumen_por_config"]:
            fila_rec = fila.copy()
            fila_rec["Opción"] = "2"
            resumen_global.append(fila_rec)
        detalles_circuitos_dfs.append(resumen_op2["detalle_por_set"])

        # --- Opción 3 ---
        opcion = 3
        descripcion_opcion = 'SET calculadas. Criterio límite de corriente en barras (SETs modulares pequeñas)'
        descripcion_opcion, opcion, wtgs_por_set_filtrada_op3 = opcion3(descripcion_opcion, opcion)
        sets_coord_op3 = np.array([[np.mean([p[1] for p in grupo]), np.mean([p[2] for p in grupo])] for grupo in wtgs_por_set_filtrada_op3])
        resumen_sets_funcion_op3 = calculos_comunes(sets_coord_op3, 0, wtgs_por_set_filtrada_op3, opcion, n_clusters_max, 0)

        base_excel_op3 = f"{descripcion_opcion}, opción {opcion}.xlsx"
        resumen_op3 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='./Software SETs/RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op3,
            archivo_final=os.path.join(r"./Software SETs/Resultados Circuitos",f"resumen_circuitos_opcion3.xlsx"),
            archivo_perdidas=os.path.join(r"./Software SETs/Pérdidas Totales",f"Pérdidas_Totales_opcion3.xlsx"),
            archivos_filtrar=[base_excel_op3]
        )
        for fila in resumen_op3["resumen_por_config"]:
            fila_rec = fila.copy()
            fila_rec["Opción"] = "3"
            resumen_global.append(fila_rec)
        detalles_circuitos_dfs.append(resumen_op3["detalle_por_set"])

        
        # ---------------- Opción 4 ----------------
        opcion = 4
        descripcion_opcion = 'Coordenadas de SETs dadas. Criterio límite de corriente en barras (SETs modulares pequeñas)'
        criterio_distancia, descripcion_opcion, opcion, wtgs_por_set_filtrada_op4 = opcion4(descripcion_opcion, opcion)

        # Reconstrucción completa (igual que tu bloque original)
        wb_sets = load_workbook("SET_Coord_prueba_300MVA.xlsx", read_only=True)
        sheet_sets = wb_sets.active
        sets_info_op4 = [list(row) for row in sheet_sets.iter_rows(values_only=True)]
        sets_coord_op4 = np.array(sets_info_op4)

        wtg_df_local = pd.read_excel(coordenadas_wtg_excel, engine="openpyxl", header=None)
        wtg_info_clean_local = [[entry[0], float(entry[1]), float(entry[2])] for entry in wtg_df_local.values.tolist()]
        wtg_coords_local = np.array([[entry[1], entry[2]] for entry in wtg_info_clean_local])

        max_wtg_per_set = math.floor(max_trafo_power / wtg_unit_power)
        wtgs_por_set_op4 = [[] for _ in range(len(sets_coord_op4))]
        set_capacities = [0] * len(sets_coord_op4)
        remaining_indices = list(range(len(wtg_coords_local)))

        while remaining_indices:
            distances = np.linalg.norm(wtg_coords_local[remaining_indices][:, np.newaxis] - sets_coord_op4, axis=2)
            new_remaining = []
            for i_idx, wtg_idx in enumerate(remaining_indices):
                sorted_sets = np.argsort(distances[i_idx])
                assigned = False
                for s in sorted_sets:
                    if set_capacities[s] < max_wtg_per_set:
                        wtgs_por_set_op4[s].append(wtg_info_clean_local[wtg_idx])
                        set_capacities[s] += 1
                        assigned = True
                        break
                if not assigned:
                    new_remaining.append(wtg_idx)
            if len(new_remaining) == len(remaining_indices):
                print("Aviso: no se pudo asignar algunos WTGs por falta de capacidad en todos los SETs.")
                break
            remaining_indices = new_remaining

        wtgs_por_set_filtrada_op4 = [sublista for sublista in wtgs_por_set_op4 if sublista]

        # Ahora sí puedes usar estas variables:
        resumen_sets_dict_op4 = calculos_comunes(sets_coord_op4, 0, wtgs_por_set_filtrada_op4, opcion, n_clusters_max, 0)


        base_excel_op4 = f"{descripcion_opcion}, opción {opcion}.xlsx"

        resumen_op4 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='./Software SETs/RESULTADOS',
            resumen_sets_funcion=resumen_sets_dict_op4,          # dict con "resumen_sets"
            archivo_final=os.path.join(r"./Software SETs/Resultados Circuitos",f"resumen_circuitos_opcion4.xlsx"),
            archivo_perdidas=os.path.join(r"./Software SETs/Pérdidas Totales",f"Pérdidas_Totales_opcion4.xlsx"),
            archivos_filtrar=[base_excel_op4]                    # nombre del .xlsx que guarda calculos_comunes para opción != 1
        )


        # Añadir al resumen global y al detalle por SET
        for fila in resumen_op4["resumen_por_config"]:
            fila_rec = fila.copy()
            fila_rec["Opción"] = "4"
            resumen_global.append(fila_rec)

        detalles_circuitos_dfs.append(resumen_op4["detalle_por_set"])



        # === CREAR EXCEL RECOPILATORIO ===
        df_resumen_global = pd.DataFrame(resumen_global)
        # Orden sugerido de columnas
        cols_orden = ["Opción","Configuración","Pérdidas en kW","Pérdidas en %",
                    "Circuitos 3 WTGs","Circuitos 2 WTGs","Nº Total Circuitos",
                    "Supply 120mm2 (m)","Supply 300mm2 (m)","Supply 630mm2 (m)","Supply earthing (m)"]
        df_resumen_global = df_resumen_global[[c for c in cols_orden if c in df_resumen_global.columns]]

        # Concatenar detalles por SET (cada df tiene columnas con el nombre_base)
        df_detalle_circuitos = pd.concat(detalles_circuitos_dfs, axis=1)

        # Guardar en un solo Excel con varias hojas
        with pd.ExcelWriter("RECOPILATORIO_Perdidas_y_Circuitos.xlsx", engine="openpyxl") as writer:
            df_resumen_global.to_excel(writer, sheet_name="Resumen Global", index=False)
            df_detalle_circuitos.to_excel(writer, sheet_name="Detalle Circuitos", index=False)

        print("Archivo 'RECOPILATORIO_Perdidas_y_Circuitos.xlsx' generado con éxito.")
        opcion = 5

    ##########COMPARATIVA RESULTADOS CAPEX (solo para la opción 5 por ahora)############
    if opcion==5:
        folder_path = "RESULTADOS"
        resumen = []
        for filename in os.listdir(folder_path):
            if filename.endswith(".xlsx"):
                file_path = os.path.join(folder_path, filename)
                try:
                    df = pd.read_excel(file_path, engine="openpyxl")
                    if "CAPEX" in df.columns:
                        # Convertir a numérico y tomar solo el primer valor no nulo
                        capex_values = df["CAPEX"].apply(lambda x: float(str(x).replace(",", "")) if pd.notnull(x) else None).dropna().tolist()
                        # Si hay valores, tomar el primero; si no, dejar vacío
                        capex_value = capex_values[0] if capex_values else None
                    else:
                        capex_value = None
                except Exception as e:
                    capex_value = None
                resumen.append({
                    "Archivo": filename,
                    "CAPEX": capex_value
                })

        # Crear DataFrame resumen
        df_resumen = pd.DataFrame(resumen)

        # Guardar en Excel
        output_file = "CAPEX_resumen_resultados.xlsx"
        df_resumen.to_excel(output_file, index=False, engine="openpyxl")

        # Aplicar formato numérico a la columna CAPEX
        wb = load_workbook(output_file)
        ws = wb.active

        # Buscar la columna CAPEX y aplicar formato
        for row in ws.iter_rows(min_row=2, min_col=2, max_col=2, max_row=ws.max_row):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0.00'
        wb.save(output_file)

# >>>> EJECUCIÓN DIRECTA
if __name__ == "__main__":
    main_set_medium_voltage()

