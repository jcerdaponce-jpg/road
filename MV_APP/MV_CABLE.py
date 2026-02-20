
import numpy as np
import math
from openpyxl import load_workbook
from collections import defaultdict
import pandas as pd
import os
from itertools import combinations
import sys
# =========================
# Opción 8: control de plots
# =========================
# True = NO dibujar (no importar Matplotlib)
# False = SÍ dibujar (importa Matplotlib y ejecuta los bloques de plotting)


NO_PLOT = True




def ensure_output_dirs(base_dir: str = '.') -> dict:
    """
    Crea (si no existen) todas las carpetas usadas por el script
    y devuelve sus rutas para reutilizarlas.

    base_dir: directorio base (por defecto, '.')

    return: dict con claves:
      - 'RESULTADOS'
      - 'RESULTADOS_WTGs' (subcarpeta: RESULTADOS/WTGs fuera de radio/SETs)
      - 'IMAGENES'
      - 'RECOPILATORIO'
    """
    base_dir = os.path.abspath(base_dir)
    rutas = {
        'RESULTADOS': os.path.join(base_dir, 'RESULTADOS'),
        'RESULTADOS_WTGs': os.path.join(base_dir, 'RESULTADOS', 'WTGs fuera de radio', 'SETs'),
        'IMAGENES': os.path.join(base_dir, 'IMAGENES'),
        'RECOPILATORIO': os.path.join(base_dir, 'RECOPILATORIO'),
        'JSON': os.path.join(base_dir, 'JSON')
    }
    os.makedirs(rutas['RESULTADOS'], exist_ok=True)
    os.makedirs(os.path.join(base_dir, 'RESULTADOS', 'WTGs fuera de radio'), exist_ok=True)
    os.makedirs(rutas['RESULTADOS_WTGs'], exist_ok=True)
    os.makedirs(rutas['IMAGENES'], exist_ok=True)
    os.makedirs(rutas['RECOPILATORIO'], exist_ok=True)
    os.makedirs(rutas['JSON'], exist_ok=True)
    return rutas


sets_coord_global = None
def main_set_medium_voltage(ncl,S_WTG,MV_VOLTAGE,wtg_list,opcion):
    """
    Flujo principal. Mantiene el orden de inputs original y añade la entrada de n_clusters_max.
    Genera Excels bajo carpetas locales y, si NO_PLOT=False, también PNGs.
    """
    # -------------------------
    # Entradas técnicas (input())
    # -------------------------
    try:
        _ncl = ncl
        n_clusters_max = int(_ncl) if _ncl else 5
        if n_clusters_max < 1:
            n_clusters_max = 1
    except Exception:
        n_clusters_max = 5

    wtg_unit_power = S_WTG * 10**3          # W
    medium_voltage_level = MV_VOLTAGE* 10**3      # V_ll
    distancia_radio_desde_SET =10500
    busbar_limitation = 2500
    max_trafo_power = 300 * 10**6  # W
    safety_factor_length = 1.08  # porcentaje aplicado a la longitud en planta, normalmente 8%

    # CAPEX (placeholders; sustitúyelos por valores de tu DB si quieres)
    cost_MV_Collector_system_300_MVA = 1      # €/MV
    cost_MV_Collector_system_10_5_km = 1      # €/MV
    cost_transformer_300_MVA = 1     # €
    cost_busbar_2500_A = 1            # €
    cost_SET = 1                    # €

    # -------------------------
    # Carpetas de salida
    # -------------------------

    _dirs = ensure_output_dirs('.')  # crea y devuelve rutas

    for d in _dirs:
        os.makedirs(d, exist_ok=True)



    # Leer filas del excel -> wtg_info: [nombre, x, y]
    wtg_info = []
    for fila in wtg_list:

        wtg_info.append(list(fila))

    n_wtg = len(wtg_info)
    pot_total_windfarm = n_wtg * wtg_unit_power  # W

    # ------------------------------------------------
    # Aux: asignación y resumen de WTGs por cada SET
    # ------------------------------------------------
    def asignar_wtgs_a_sets(sets_coord, wtg_info_local, usar_indice_desde_1=True):
        resultados = {}
        wtg_en_sets = defaultdict(list)
        asignaciones = {}

        for fila in wtg_info_local:
            id_wtg, x_wtg, y_wtg = fila[0], fila[1], fila[2]
            min_dist = float('inf')
            set_idx_min = None
            distancias = []
            for idx, (x_set, y_set) in enumerate(sets_coord):
                dist = math.hypot(x_wtg - x_set, y_wtg - y_set)
                distancias.append(dist)
                if dist < min_dist:
                    min_dist = dist
                    set_idx_min = idx
            dist_min_km = min(distancias) / 1000.0
            dist_max_km = max(distancias) / 1000.0
            resultados[id_wtg] = {"dist_min": dist_min_km, "dist_max": dist_max_km}
            indice_reportado = set_idx_min + 1 if usar_indice_desde_1 else set_idx_min
            wtg_en_sets[indice_reportado].append(id_wtg)
            asignaciones[id_wtg] = {"set": indice_reportado, "dist": min_dist}
        return wtg_en_sets, asignaciones, resultados

    def resumen_por_set_completo(sets_coord, wtg_info_local, wtg_en_sets, usar_vecino_mas_cercano=True):
        wtg_dict = {fila[0]: (fila[1], fila[2]) for fila in wtg_info_local}
        resumen = []
        for set_idx, lista_wtgs in wtg_en_sets.items():
            if not lista_wtgs:
                resumen.append({
                    "SET": set_idx,
                    "WTGs": 0,
                    "Dist_Min": None,
                    "Dist_Max": None,
                    "Dist_Media_WTGs": None
                })
                continue
            x_set, y_set = sets_coord[set_idx - 1]
            distancias_set = [math.hypot(wtg_dict[w][0] - x_set, wtg_dict[w][1] - y_set) for w in lista_wtgs]
            dist_min = min(distancias_set) / 1000.0
            dist_max = max(distancias_set) / 1000.0

            if len(lista_wtgs) < 2:
                dist_media_wtgs = None
            else:
                if usar_vecino_mas_cercano:
                    distancias_minimas = []
                    for wtg_id in lista_wtgs:
                        distancias_otras = [
                            math.hypot(wtg_dict[wtg_id][0] - wtg_dict[otro][0],
                                       wtg_dict[wtg_id][1] - wtg_dict[otro][1])
                            for otro in lista_wtgs if otro != wtg_id
                        ]
                        distancias_minimas.append(min(distancias_otras))
                    dist_media_wtgs = (sum(distancias_minimas) / len(distancias_minimas)) / 1000.0
                else:
                    distancias_parejas = [
                        math.hypot(wtg_dict[a][0] - wtg_dict[b][0], wtg_dict[a][1] - wtg_dict[b][1])
                        for a, b in combinations(lista_wtgs, 2)
                    ]
                    dist_media_wtgs = (sum(distancias_parejas) / len(distancias_parejas)) / 1000.0

            resumen.append({
                "SET": set_idx,
                "WTGs": len(lista_wtgs),
                "Dist_Min": dist_min,
                "Dist_Max": dist_max,
                "Dist_Media_WTGs": dist_media_wtgs
            })
        return resumen

    # ------------------------------------------------
    # Pérdidas por SET (tu lógica original)
    # ------------------------------------------------
    def calculo_perdidas(resumen_sets_funcion, safety_factor_length_local):
        # Acepta dict {"resumen_sets": [...]} o lista
        if isinstance(resumen_sets_funcion, dict) and "resumen_sets" in resumen_sets_funcion:
            resumen_sets = resumen_sets_funcion["resumen_sets"]
        else:
            resumen_sets = resumen_sets_funcion

        # Corrientes (PF=0.95)
        current_1WTG = (((wtg_unit_power / 1000) / 0.95) / (math.sqrt(3) * medium_voltage_level / 1000))
        current_2WTG = ((((2 * wtg_unit_power) / 1000) / 0.95) / (math.sqrt(3) * medium_voltage_level / 1000))
        current_3WTG = ((((3 * wtg_unit_power) / 1000) / 0.95) / (math.sqrt(3) * medium_voltage_level / 1000))

        # Resistencias (Ω/km)
        resist_120mm2 = 0.325
        resist_300mm2 = 0.130
        resist_630mm2 = 0.063

        perdidas_3WTG_por_set = []
        perdidas_2WTG_por_set = []
        perdidas_1WTG_por_set = []
        supply_120mm2_3WTGs = []
        supply_300mm2_3WTGs = []
        supply_630mm2_3WTGs = []
        supply_120mm2_2WTGs = []
        supply_300mm2_2WTGs = []
        supply_630mm2_1WTGs = []

        for r in resumen_sets:
            dist_media = r["Dist_Media_WTGs"]
            dist_min = r["Dist_Min"]
            dist_max = r["Dist_Max"]
            if dist_media is None or dist_min is None or dist_max is None:
                perdidas_3WTG_por_set.append(0.0)
                perdidas_2WTG_por_set.append(0.0)
                perdidas_1WTG_por_set.append(0.0)
                supply_120mm2_3WTGs.append(0.0)
                supply_300mm2_3WTGs.append(0.0)
                supply_630mm2_3WTGs.append(0.0)
                supply_120mm2_2WTGs.append(0.0)
                supply_300mm2_2WTGs.append(0.0)
                supply_630mm2_1WTGs.append(0.0)
                continue

            d_btw = dist_media * safety_factor_length_local
            d_mean = ((dist_max + dist_min) / 2) * safety_factor_length_local

            # 3 WTGs
            p_1_3 = 3 * (current_1WTG ** 2) * (d_btw * resist_120mm2)
            s_120_3 = d_btw * 3
            p_2_3 = 3 * (current_2WTG ** 2) * (d_btw * resist_300mm2)
            s_300_3 = d_btw * 3
            p_3_3 = 3 * (current_3WTG ** 2) * (d_mean * resist_630mm2)
            s_630_3 = d_mean * 3
            perdidas_3WTG_por_set.append(p_1_3 + p_2_3 + p_3_3)
            supply_120mm2_3WTGs.append(s_120_3)
            supply_300mm2_3WTGs.append(s_300_3)
            supply_630mm2_3WTGs.append(s_630_3)

            # 2 WTGs
            p_1_2 = 3 * (current_1WTG ** 2) * (d_btw * resist_120mm2)
            s_120_2 = d_btw * 3
            p_2_2 = 3 * (current_2WTG ** 2) * (d_mean * resist_630mm2)
            s_300_2 = d_mean * 3  # tu lógica original
            perdidas_2WTG_por_set.append(p_1_2 + p_2_2)
            supply_120mm2_2WTGs.append(s_120_2)
            supply_300mm2_2WTGs.append(s_300_2)

            # 1 WTG
            p_1_1 = 3 * (current_1WTG ** 2) * (d_mean * resist_630mm2)
            s_630_1 = d_mean * 3
            perdidas_1WTG_por_set.append(p_1_1)
            supply_630mm2_1WTGs.append(s_630_1)

        return (
            perdidas_3WTG_por_set,
            perdidas_2WTG_por_set,
            perdidas_1WTG_por_set,
            supply_120mm2_3WTGs,
            supply_300mm2_3WTGs,
            supply_630mm2_3WTGs,
            supply_120mm2_2WTGs,
            supply_300mm2_2WTGs,
            supply_630mm2_1WTGs
        )

    # ------------------------------------------------
    # Generación y combinación de Excels por configuración
    # ------------------------------------------------
    def procesar_excels(
        wtg_opcion1, wtg_opcion2, wtg_opcion3, wtg_opcion4,
        carpeta_resultados, resumen_sets_funcion,
        archivo_final="RESULTADOS/resultados_circuitos.xlsx",
        archivo_perdidas="RESULTADOS/Pérdidas Totales_in kW.xlsx",
        archivos_filtrar=None
    ):
        # Aceptar dict {"resumen_sets": [...]} o lista
        if isinstance(resumen_sets_funcion, dict) and "resumen_sets" in resumen_sets_funcion:
            resumen_sets_funcion = resumen_sets_funcion["resumen_sets"]
        if not isinstance(resumen_sets_funcion, list):
            raise ValueError("Se esperaba una lista de diccionarios en resumen_sets_funcion.")

        (
            p3_list, p2_list, p1_list,
            s120_3_list, s300_3_list, s630_3_list,
            s120_2_list, s300_2_list, s630_1_list
        ) = calculo_perdidas(resumen_sets_funcion, safety_factor_length)

        # Selección de archivos a procesar
        if archivos_filtrar and len(archivos_filtrar) > 0:
            archivos = archivos_filtrar
        else:
            archivos = [f for f in os.listdir(carpeta_resultados) if f.endswith('.xlsx')]

        lista_cantidad = []
        lista_circuitos_3 = []
        lista_circuitos_2 = []
        lista_cantidad_120_mm2 = []
        lista_cantidad_300_mm2 = []
        lista_cantidad_630_mm2 = []
        dataframes = []
        solo_perdidas_pruebas = []

        for archivo in archivos:
            ruta_completa = os.path.join(carpeta_resultados, archivo)
            try:
                df = pd.read_excel(ruta_completa)
            except FileNotFoundError:
                ruta_alt = os.path.join(carpeta_resultados, "WTGs fuera de radio", archivo)
                df = pd.read_excel(ruta_alt)
            if 'Cantidad por SET' not in df.columns:
                print(f"⚠ El archivo '{archivo}' no tiene la columna 'Cantidad por SET'. Se omite.")
                continue

            cantidad = df['Cantidad por SET']

            # Coherencia: nº filas vs nº SETs en pérdidas
            if len(cantidad) != len(p3_list):
                raise ValueError(
                    f"Desajuste de longitudes para '{archivo}': "
                    f"{len(cantidad)} filas (SETs en Excel) vs {len(p3_list)} SETs en pérdidas."
                    f"Asegúrate de pasar a 'archivos_filtrar' el archivo correcto para esta configuración."
                )

            # Derivar nº circuitos 3/2 WTGs (regla 0.333/0.667)
            n_circuitos_3_wtg = cantidad / 3
            n_circuitos_2_wtg = pd.Series([0] * len(cantidad))
            parte_decimal = n_circuitos_3_wtg % 1
            parte_decimal_redondeada = parte_decimal.round(3)

            for i in range(len(cantidad)):
                if parte_decimal_redondeada[i] == 0.333:
                    n_circuitos_3_wtg[i] = (cantidad[i] - 4) / 3
                    n_circuitos_2_wtg[i] = 2
                elif parte_decimal_redondeada[i] == 0.667:
                    n_circuitos_3_wtg[i] = (cantidad[i] - 2) / 3
                    n_circuitos_2_wtg[i] = 1
                else:
                    n_circuitos_3_wtg[i] = int(n_circuitos_3_wtg[i])
                    n_circuitos_2_wtg[i] = 0

            # Pérdidas y supply por SET
            perdidas_totales_SET = []
            supply_120_total = []
            supply_300_total = []
            supply_630_total = []

            for i in range(len(cantidad)):
                perdidas_3 = n_circuitos_3_wtg[i] * p3_list[i]
                perdidas_2 = n_circuitos_2_wtg[i] * p2_list[i]
                perdidas_kW = (perdidas_3 + perdidas_2) / 1000.0
                perdidas_totales_SET.append(perdidas_kW)

                s120 = (n_circuitos_3_wtg[i] * s120_3_list[i]) * 3 + (n_circuitos_2_wtg[i] * s120_2_list[i]) * 3
                s300 = (n_circuitos_3_wtg[i] * s300_3_list[i]) * 3 + (n_circuitos_2_wtg[i] * s300_2_list[i]) * 3
                s630 = (n_circuitos_3_wtg[i] * s630_3_list[i]) * 3 + (n_circuitos_2_wtg[i] * s630_1_list[i]) * 3
                supply_120_total.append(s120)
                supply_300_total.append(s300)
                supply_630_total.append(s630)

            lista_cantidad.append(cantidad.sum())
            lista_circuitos_3.append(n_circuitos_3_wtg.sum())
            lista_circuitos_2.append(n_circuitos_2_wtg.sum())
            lista_cantidad_120_mm2.append(sum(supply_120_total))
            lista_cantidad_300_mm2.append(sum(supply_300_total))
            lista_cantidad_630_mm2.append(sum(supply_630_total))

            nombre_base = os.path.splitext(archivo)[0]
            df_resultado = pd.DataFrame({
                nombre_base + '_Cir_3WTG': n_circuitos_3_wtg,
                nombre_base + '_Cir_2WTG': n_circuitos_2_wtg,
                nombre_base + '_Total_Cantidad': cantidad,
                nombre_base + '_Pérdidas (kW)': perdidas_totales_SET,
                nombre_base + '_Supply 120mm2 (km)': supply_120_total,
                nombre_base + '_Supply 300mm2 (km)': supply_300_total,
                nombre_base + '_Supply 630mm2 (km)': supply_630_total,
            })
            dataframes.append(df_resultado)

            total_perdidas = sum(perdidas_totales_SET)
            df_solo_perdidas = pd.DataFrame({nombre_base + '_PÉRDIDAS TOTALES': [total_perdidas]})
            solo_perdidas_pruebas.append(df_solo_perdidas)

        df_final = pd.concat(dataframes, axis=1) if dataframes else pd.DataFrame()
        df_solo_perdidas_final = pd.concat(solo_perdidas_pruebas, axis=1).T.reset_index() if solo_perdidas_pruebas else pd.DataFrame(columns=['Configuración','Pérdidas en kW'])

        if not df_solo_perdidas_final.empty:
            df_solo_perdidas_final.columns = ['Configuración', 'Pérdidas en kW']
            df_solo_perdidas_final['Pérdidas en %'] = ((df_solo_perdidas_final['Pérdidas en kW'] * 1000) / pot_total_windfarm) * 100
            df_solo_perdidas_final['Circuitos de 3 WTGs'] = lista_circuitos_3
            df_solo_perdidas_final['Circuitos de 2 WTGs'] = lista_circuitos_2
            df_solo_perdidas_final['Número Total Circuitos'] = df_solo_perdidas_final['Circuitos de 3 WTGs'] + df_solo_perdidas_final['Circuitos de 2 WTGs']
            df_solo_perdidas_final['Supply 120mm2 cable (km)'] = lista_cantidad_120_mm2
            df_solo_perdidas_final['Supply 300mm2 cable (km)'] = lista_cantidad_300_mm2
            df_solo_perdidas_final['Supply 630mm2 cable (km)'] = lista_cantidad_630_mm2
            df_solo_perdidas_final['Supply earthing cable (km)'] = (
                df_solo_perdidas_final['Supply 120mm2 cable (km)'] +
                df_solo_perdidas_final['Supply 300mm2 cable (km)'] +
                df_solo_perdidas_final['Supply 630mm2 cable (km)']
            )

        # Guardado
        if not df_final.empty:
            os.makedirs(os.path.dirname(archivo_final), exist_ok=True)
            df_final.to_excel(archivo_final, index=False)
            #print(f"Archivo Excel generado: {archivo_final}")
        if not df_solo_perdidas_final.empty:
            os.makedirs(os.path.dirname(archivo_perdidas), exist_ok=True)
            df_solo_perdidas_final.to_excel(archivo_perdidas, index=False)
            #print(f"Archivo Excel generado: {archivo_perdidas}")

        resumen_config = []
        if not df_solo_perdidas_final.empty:
            for _, fila in df_solo_perdidas_final.iterrows():
                resumen_config.append({
                    "Configuración": fila["Configuración"],
                    "Pérdidas en kW": fila["Pérdidas en kW"],
                    "Pérdidas en %": fila["Pérdidas en %"],
                    "Circuitos 3 WTGs": fila["Circuitos de 3 WTGs"],
                    "Circuitos 2 WTGs": fila["Circuitos de 2 WTGs"],
                    "Nº Total Circuitos": fila["Número Total Circuitos"],
                    "Supply 120mm2 (km)": fila["Supply 120mm2 cable (km)"],
                    "Supply 300mm2 (km)": fila["Supply 300mm2 cable (km)"],
                    "Supply 630mm2 (km)": fila["Supply 630mm2 cable (km)"],
                    "Supply earthing (km)": fila["Supply earthing cable (km)"]
                })

        detalle_por_set = df_final.copy() if not df_final.empty else pd.DataFrame()
        return {
            "archivo_resultados": archivo_final,
            "archivo_perdidas": archivo_perdidas,
            "resumen_por_config": resumen_config,
            "detalle_por_set": detalle_por_set
        }

    # ------------------------------------------------
    # Comunes (dimensionamiento, CAPEX, Excel por SET)
    # ------------------------------------------------
    def calculos_comunes(sets_coord, descripcion_opcion, criterio_distancia, wtgs_por_set_filtrada, opcion_local, n_clusters_max_local, wtg_fuera_radio):
        wtg_en_sets, asignaciones, resultados = asignar_wtgs_a_sets(sets_coord, wtg_info, usar_indice_desde_1=True)
        resumen_sets_funcion_local = resumen_por_set_completo(sets_coord, wtg_info, wtg_en_sets, usar_vecino_mas_cercano=True)

        # Alinear longitudes si hay discrepancias
        if len(sets_coord) != len(wtgs_por_set_filtrada):
            mask_no_vacio = [len(sublista) > 0 for sublista in wtgs_por_set_filtrada]
            if all(mask_no_vacio):
                sets_coord = np.array(sets_coord, dtype=float)[:len(wtgs_por_set_filtrada)]
            else:
                sets_coord = np.array(sets_coord, dtype=float)[mask_no_vacio]
                wtgs_por_set_filtrada = [s for s in wtgs_por_set_filtrada if len(s) > 0]

        n_sets = len(sets_coord)

        conteo_por_set = [len(lista_wtgs) for lista_wtgs in wtgs_por_set_filtrada]
        lista_cantidad = [cantidad for cantidad in conteo_por_set]
        pot_per_SET_MW = [valor * wtg_unit_power / (10**6) for valor in lista_cantidad]

        number_busbar = [(((valor * 10**6) / (math.sqrt(3) * medium_voltage_level)) / busbar_limitation) for valor in pot_per_SET_MW]
        number_busbar_redondeados_arriba = [math.ceil(valor) for valor in number_busbar]
        n_totales_busbar = sum(number_busbar_redondeados_arriba)

        number_transformers = [((valor * 10**6) / max_trafo_power) for valor in pot_per_SET_MW]
        number_transformers_redondeados_arriba = [math.ceil(valor) for valor in number_transformers]
        n_totales_transformers = sum(number_transformers_redondeados_arriba)

        if criterio_distancia == 1:
            CAPEX = cost_MV_Collector_system_10_5_km * n_wtg * wtg_unit_power +                     n_totales_busbar * cost_busbar_2500_A +                     n_totales_transformers * cost_transformer_300_MVA +                     n_sets * cost_SET
        else:
            CAPEX = cost_MV_Collector_system_300_MVA * n_wtg * wtg_unit_power +                     n_totales_busbar * cost_busbar_2500_A +                     n_totales_transformers * cost_transformer_300_MVA +                     n_sets * cost_SET

        sets_coord_str = [f"[{x[0]:.6f}, {x[1]:.6f}]" for x in sets_coord]
        capex_column = [""] * len(sets_coord_str)
        capex_formateado = f"{CAPEX:,.2f}"
        if len(capex_column) >= 1:
            capex_column[0] = capex_formateado

        coord_x = [coord[0] for coord in sets_coord]
        coord_y = [coord[1] for coord in sets_coord]
        df = pd.DataFrame({
            "Coord X SET": coord_x,
            "Coord Y SET": coord_y,
            "Cantidad por SET": lista_cantidad,
            "Potencia por SET (MW)": pot_per_SET_MW,
            "Busbars por SET": number_busbar_redondeados_arriba,
            "Transformadores por SET": number_transformers_redondeados_arriba,
            "CAPEX": capex_column
        })

        # Excel naming EXACTO igual al esperado por otros bloques
        if opcion_local == 5:
            pass
        elif opcion_local != 1 and wtg_fuera_radio == 0:
            excel_filename = "RESULTADOS/" + str(descripcion_opcion) + ", opción " + str(opcion_local) + ".xlsx"
        elif opcion_local != 1 and wtg_fuera_radio == 1:
            excel_filename = "RESULTADOS/WTGs fuera de radio/" + str(descripcion_opcion) + ", opción " + str(opcion_local) + ".xlsx"
        elif opcion_local == 1 and wtg_fuera_radio == 0:
            excel_filename = "RESULTADOS/" + str(descripcion_opcion) + " " + str(n_clusters_max_local) + " SETs, opción " + str(opcion_local) + ".xlsx"
        else:  # opcion_local == 1 and wtg_fuera_radio == 1
            excel_filename = "RESULTADOS/WTGs fuera de radio/" + str(descripcion_opcion) + " " + str(n_clusters_max_local) + " SETs, opción " + str(opcion_local) + ".xlsx"

        if opcion_local != 5:
            base_dir = os.path.dirname(excel_filename)
            if base_dir and not os.path.exists(base_dir):
                os.makedirs(base_dir, exist_ok=True)
            df.to_excel(excel_filename, index=False)
            #print(f"Archivo Excel generado: {excel_filename}")

        return {
            "wtgs_por_set": wtgs_por_set_filtrada,
            "cantidad_por_set": lista_cantidad,
            "potencia_por_set_MW": pot_per_SET_MW,
            "busbars_por_set": number_busbar_redondeados_arriba,
            "transformadores_por_set": number_transformers_redondeados_arriba,
            "resumen_sets": resumen_sets_funcion_local
        }

    # ------------------------------------------------
    # Con criterio de distancia (radio) + plotting opcional
    # ------------------------------------------------
    def calculos_comunes_criterio_distancia(opcion_local, n_clusters_max_local, descripcion_opcion,
                                            criterio_distancia, wtg_info_local, sets_coord, distancia_radio_desde_SET_local):
        # Construir grupos por radio
        wtg_en_sets = defaultdict(list)
        for idx, set_coord in enumerate(sets_coord):
            for fila in wtg_info_local:
                id_wtg, x, y = fila[0], fila[1], fila[2]
                distancia = math.sqrt((x - set_coord[0])**2 + (y - set_coord[1])**2)
                if 0 < distancia < distancia_radio_desde_SET_local:
                    wtg_en_sets[id_wtg].append(idx + 1)

        # SET más cercano / fuera de radio
        wtg_mas_cercano = []
        wtgs_fuera_de_todas = []
        for wtg in wtg_info_local:
            id_wtg, x, y = wtg[0], wtg[1], wtg[2]
            distancia_min = float('inf')
            set_mas_cercano = None
            for idx, set_coord in enumerate(sets_coord):
                distancia = math.sqrt((x - set_coord[0])**2 + (y - set_coord[1])**2)
                if distancia < distancia_min:
                    distancia_min = distancia
                    set_mas_cercano = idx
            # registrar set más cercano para cada WTG
            wtg_mas_cercano.append((id_wtg, set_mas_cercano, distancia_min))
            # marcar fuera de radio si aplica
            if distancia_min > distancia_radio_desde_SET_local:
                coord_set_cercano = (float(sets_coord[set_mas_cercano][0]), float(sets_coord[set_mas_cercano][1]))
                wtgs_fuera_de_todas.append((id_wtg, coord_set_cercano))
            wtg_fuera_radio = 1 if distancia_min > distancia_radio_desde_SET_local else 0

        # Crear lista por SET (sin vacíos)
        wtgs_por_set = [[] for _ in range(len(sets_coord))]
        for wtg_id, set_idx, dist in wtg_mas_cercano:
            wtgs_por_set[set_idx].append(wtg_id)

        mask_no_vacio = [len(sublista) > 0 for sublista in wtgs_por_set]
        sets_coord_filtrada = np.array(sets_coord)[mask_no_vacio]
        wtgs_por_set_filtrada = [sublista for sublista in wtgs_por_set if sublista]

        # Plot opcional (NO_PLOT)
        if not NO_PLOT:
            import matplotlib.pyplot as plt
            import matplotlib as mpl
            from matplotlib.patches import Circle, Patch
            import matplotlib.patches as patches

            fig, ax = plt.subplots(figsize=(10, 8))
            wtg_coord_dict = {wtg[0]: (wtg[1], wtg[2]) for wtg in wtg_info_local}
            colors = plt.get_cmap('tab10', len(wtgs_por_set_filtrada))

            scatter_wtg_handles = []
            for i, grupo in enumerate(wtgs_por_set_filtrada):
                wtg_xs, wtg_ys = [], []
                for wtg_id in grupo:
                    if wtg_id in wtg_coord_dict:
                        xx, yy = wtg_coord_dict[wtg_id]
                        wtg_xs.append(xx); wtg_ys.append(yy)
                scatter_wtg = ax.scatter(wtg_xs, wtg_ys, label=f'SET {i+1} ({len(grupo)} WTGs)', s=30, color=colors(i))
                scatter_wtg_handles.append(scatter_wtg)

            scatter_sets = ax.scatter(sets_coord_filtrada[:, 0], sets_coord_filtrada[:, 1], color='black', marker='X', s=100, label='SETs')
            ax.legend(handles=scatter_wtg_handles + [scatter_sets])
            ax.set_xlabel('Coordenada X'); ax.set_ylabel('Coordenada Y')
            ax.set_title("Opción" + str(opcion_local) + ". " + str(descripcion_opcion) + ".")
            ax.grid(True)
            plt.axis('equal')
            plt.tight_layout()

            if wtg_fuera_radio == 1:
                for x_set, y_set in sets_coord:
                    circle = Circle((x_set, y_set), distancia_radio_desde_SET_local, color='blue', fill=False, linestyle='--', linewidth=2, alpha=0.7)
                    ax.add_patch(circle)
                legend_circle = Patch(edgecolor='blue', facecolor='none', linestyle='--', linewidth=2,
                                      label=f'Radius {distancia_radio_desde_SET_local/1000} km')
                ax.legend(handles=scatter_wtg_handles + [scatter_sets, legend_circle])

                xlim = ax.get_xlim(); ylim = ax.get_ylim()
                ancho = (xlim[1] - xlim[0]) * 0.25
                alto = (ylim[1] - ylim[0]) * 0.15
                x_rect = xlim[1] - ancho - (xlim[1] - xlim[0]) * 0.002
                y_rect = ylim[0] + (ylim[1] - ylim[0]) * 0.01
                rect = patches.FancyBboxPatch((x_rect, y_rect), ancho, alto,
                                              boxstyle="round,pad=0.3",
                                              linewidth=2, edgecolor='red',
                                              facecolor='lightyellow', alpha=0.9)
                ax.add_patch(rect)

                if opcion_local == 2:
                    plt.savefig("IMAGENES/" + str(descripcion_opcion) + ", opción " + str(opcion_local) + ".png", dpi=300)
                if opcion_local == 1:
                    plt.savefig("IMAGENES/" + str(descripcion_opcion) + " " + str(n_clusters_max_local) + " SETs, opción " + str(opcion_local) + " .png", dpi=300)

        # Excel de WTGs fuera de radio por SET (sin plotting)
        if len(wtgs_por_set_filtrada) > 0:
            data = []
            for set_name, items in [(f"SET_{i+1}", len(gr)) for i, gr in enumerate(wtgs_por_set_filtrada)]:
                row = [set_name, items]
                data.append(row)
            headers = ['SET', 'Item1']
            df_out = pd.DataFrame(data, columns=headers)
            divisor = 2
            df_out['Nºcircuitos de 2 WTGs'] = df_out['Item1'].astype(int) // divisor
            df_out['Nºcircuitos de 1 WTG'] = df_out['Item1'].astype(int) % divisor

            if opcion_local == 5:
                pass
            elif opcion_local != 1:
                excel_filename = "RESULTADOS/WTGs fuera de radio/SETs/" + str(descripcion_opcion) + ", opción " + str(opcion_local) + "_WTGs per SETS.xlsx"
                os.makedirs(os.path.dirname(excel_filename), exist_ok=True)
                df_out.to_excel(excel_filename, index=False)
                #print(f"Archivo Excel generado: {excel_filename}")
            else:
                excel_filename = "RESULTADOS/WTGs fuera de radio/SETs/" + str(descripcion_opcion) + " " + str(n_clusters_max_local) + "_WTGs per SETS, opción " + str(opcion_local) + "PRUEBA SETS.xlsx"
                os.makedirs(os.path.dirname(excel_filename), exist_ok=True)
                df_out.to_excel(excel_filename, index=False)
                #print(f"Archivo Excel generado: {excel_filename}")

        # Llamada a comunes (dimensionamiento + Excel principal por SET)
        resumen_sets_funcion = calculos_comunes(sets_coord_filtrada, descripcion_opcion, criterio_distancia, wtgs_por_set_filtrada, opcion_local, n_clusters_max_local, wtg_fuera_radio)
        return resumen_sets_funcion, opcion_local, n_clusters_max_local, wtg_fuera_radio, wtgs_por_set_filtrada


    # ------------------------------------------------
    # Opción 1: SETs calculadas por KMeans (SIN barrido)
    # ------------------------------------------------

    def opcion1(descripcion_opcion, opcion_local, n_clusters_max_local):
        from sklearn.cluster import KMeans
        global sets_coord_global
        coordenadas = np.array([[x[1], x[2]] for x in wtg_info])
        kmeans = KMeans(n_clusters_max_local, random_state=42)
        kmeans.fit(coordenadas)
        sets_coord = kmeans.cluster_centers_
        sets_coord_global = sets_coord
        criterio_distancia = 1
        resumen_sets, opcion_local, n_clusters_max_local, wtg_fuera_radio, wtgs_por_set_filtrada = calculos_comunes_criterio_distancia(
            opcion_local, n_clusters_max_local, descripcion_opcion, criterio_distancia, wtg_info, sets_coord, distancia_radio_desde_SET
        )
        return criterio_distancia, descripcion_opcion, opcion_local, resumen_sets, wtg_fuera_radio, wtgs_por_set_filtrada

    # ------------------------------------------------
    # Opción 2: SETs dadas (Excel) con criterio distancia
    # ------------------------------------------------
    def opcion2(descripcion_opcion, opcion_local):
        n_clusters_max_local = 0
        criterio_distancia = 1
        wb_sets = load_workbook("SET_Coord_test_run.xlsx", read_only=True)
        sheet_sets = wb_sets.active
        sets_info = []
        for fila in sheet_sets.iter_rows(values_only=True):
            sets_info.append(list(fila))
        sets_coord = np.array(sets_info)
        resumen_sets, opcion_local, n_clusters_max_local, wtg_fuera_radio, wtgs_por_set_filtrada = calculos_comunes_criterio_distancia(
            opcion_local, n_clusters_max_local, descripcion_opcion, criterio_distancia, wtg_info, sets_coord, distancia_radio_desde_SET
        )

        return criterio_distancia, resumen_sets, descripcion_opcion, opcion_local, wtgs_por_set_filtrada

    # ------------------------------------------------
    # Opción 3: SETs modulares pequeñas (límite de corriente/trafo)
    # ------------------------------------------------
    def opcion3(descripcion_opcion, opcion_local):
        criterio_distancia = 0
        n_clusters_max_local = 0

        # nº clusters inicial (ceil por potencia)
        num_clusters = math.ceil(n_wtg * wtg_unit_power / max_trafo_power)
        max_WTG_per_SET = math.floor(max_trafo_power / wtg_unit_power)

        df = pd.wtg_list()
        wtg_info_clean = [[entry[0], float(entry[1]), float(entry[2])] for entry in df.values.tolist()]
        coords = np.array([[entry[1], entry[2]] for entry in wtg_info_clean])

        from sklearn.cluster import KMeans
        kmeans = KMeans(n_clusters=num_clusters, random_state=0)
        kmeans.fit(coords)
        initial_centroids = kmeans.cluster_centers_

        assigned_groups = [[] for _ in range(num_clusters)]
        remaining_indices = list(range(len(coords)))

        while remaining_indices:
            distances = np.linalg.norm(coords[remaining_indices][:, np.newaxis] - initial_centroids, axis=2)
            new_remaining_indices = []
            for i, idx in enumerate(remaining_indices):
                sorted_centroids = np.argsort(distances[i])
                assigned = False
                for c in sorted_centroids:
                    if len(assigned_groups[c]) < max_WTG_per_SET:
                        assigned_groups[c].append(wtg_info_clean[idx])
                        assigned = True
                        break
                if not assigned:
                    new_remaining_indices.append(idx)
            if len(new_remaining_indices) == len(remaining_indices):
                print("No se pudo asignar algunos WTGs debido a la falta de espacio en todos los grupos.")
                break
            remaining_indices = new_remaining_indices

        new_centroids = []
        for grupo in assigned_groups:
            xs = [p[1] for p in grupo]
            ys = [p[2] for p in grupo]
            new_centroids.append((np.mean(xs), np.mean(ys)))
        new_centroids = np.array(new_centroids)

        # Plot opcional
        if not NO_PLOT:
            import matplotlib.pyplot as plt
            import matplotlib.patches as patches

            fig, ax = plt.subplots(figsize=(10, 8))
            for i, grupo in enumerate(assigned_groups):
                xs = [p[1] for p in grupo]
                ys = [p[2] for p in grupo]
                ax.scatter(xs, ys, label=f'SET {i+1} ({len(grupo)} WTGs)', s=30)

            xlim = ax.get_xlim(); ylim = ax.get_ylim()
            ancho = (xlim[1] - xlim[0]) * 0.25
            alto = (ylim[1] - ylim[0]) * 0.15
            x_rect = xlim[1] - ancho - (xlim[1] - xlim[0]) * 0.002
            y_rect = ylim[0] + (ylim[1] - ylim[0]) * 0.01
            rect = patches.FancyBboxPatch((x_rect, y_rect), ancho, alto,
                                          boxstyle="round,pad=0.3",
                                          linewidth=2, edgecolor='red',
                                          facecolor='lightyellow', alpha=0.9)
            ax.add_patch(rect)

            ax.scatter(new_centroids[:, 0], new_centroids[:, 1], color='black', marker='X', s=100, label='SETs')
            ax.set_title("Opción" + str(opcion_local) + ". " + str(descripcion_opcion) + ".")
            ax.grid(True); ax.legend()
            ax.set_xlabel('Coordenada X'); ax.set_ylabel('Coordenada Y')
            plt.tight_layout()
            plt.savefig("IMAGENES/" + str(descripcion_opcion) + ", opción " + str(opcion_local) + ".png", dpi=300)

        sets_coord = new_centroids
        wtgs_por_set = assigned_groups
        mask_no_vacio = [len(sublista) > 0 for sublista in wtgs_por_set]
        sets_coord_filtrada = np.array(sets_coord)[mask_no_vacio]
        wtgs_por_set_filtrada = [sublista for sublista in wtgs_por_set if sublista]

        resumen_sets_funcion = calculos_comunes(sets_coord_filtrada, descripcion_opcion, 0, wtgs_por_set_filtrada, opcion_local, n_clusters_max, 0)
        return descripcion_opcion, opcion_local, wtgs_por_set_filtrada

    # ------------------------------------------------
    # Opción 4: SETs dadas + modular pequeñas
    # ------------------------------------------------
    def opcion4(descripcion_opcion, opcion_local):
        max_wtg_per_set = math.floor(max_trafo_power / wtg_unit_power)
        criterio_distancia = 0
        n_clusters_max_local = 0

        wb = load_workbook("SET_Coord_300MVA_test_run.xlsx", read_only=True)
        sheet = wb.active
        sets_info = [list(row) for row in sheet.iter_rows(values_only=True)]
        sets_coords = np.array(sets_info)

        wtg_df = pd.wtg_list()
        wtg_info_clean = [[entry[0], float(entry[1]), float(entry[2])] for entry in wtg_df.values.tolist()]
        wtg_coords = np.array([[entry[1], entry[2]] for entry in wtg_info_clean])

        wtgs_por_set = [[] for _ in range(len(sets_coords))]
        set_capacities = [0] * len(sets_coords)
        remaining_indices = list(range(len(wtg_coords)))

        while remaining_indices:
            distances = np.linalg.norm(wtg_coords[remaining_indices][:, np.newaxis] - sets_coords, axis=2)
            new_remaining = []
            for i_idx, wtg_idx in enumerate(remaining_indices):
                sorted_sets = np.argsort(distances[i_idx])
                assigned = False
                for s in sorted_sets:
                    if set_capacities[s] < max_wtg_per_set:
                        wtgs_por_set[s].append(wtg_info_clean[wtg_idx])
                        set_capacities[s] += 1
                        assigned = True
                        break
                if not assigned:
                    new_remaining.append(wtg_idx)
            if len(new_remaining) == len(remaining_indices):
                print("No se pudo asignar algunos WTGs debido a falta de capacidad.")
                break
            remaining_indices = new_remaining

        # Plot opcional
        if not NO_PLOT:
            import matplotlib.pyplot as plt
            import matplotlib as mpl

            fig, ax = plt.subplots(figsize=(10, 8))
            colors = plt.get_cmap('tab10', len(wtgs_por_set))
            for i, grupo in enumerate(wtgs_por_set):
                xs = [p[1] for p in grupo]
                ys = [p[2] for p in grupo]
                ax.scatter(xs, ys, label=f'SET {i+1} ({len(grupo)} WTGs)', s=30, color=colors(i))
            ax.scatter(sets_coords[:, 0], sets_coords[:, 1], color='black', marker='X', s=100, label='SETs')
            ax.set_title("Opción" + str(opcion_local) + ". " + str(descripcion_opcion) + ".")
            ax.grid(True); ax.legend()
            ax.set_xlabel('Coordenada X'); ax.set_ylabel('Coordenada Y')
            plt.tight_layout()
            plt.savefig("IMAGENES/" + str(descripcion_opcion) + " ,opción " + str(opcion_local) + ".png", dpi=300)

        wtgs_por_set_filtrada = [sublista for sublista in wtgs_por_set if sublista]
        mask_no_vacio = [len(grupo) > 0 for grupo in wtgs_por_set]
        sets_coords_filtrada = np.array(sets_coords, dtype=float)[mask_no_vacio]
        wtgs_por_set_filtrada = [grupo for grupo, keep in zip(wtgs_por_set, mask_no_vacio) if keep]

        resumen_sets_funcion = calculos_comunes(sets_coords_filtrada, descripcion_opcion, 0, wtgs_por_set_filtrada, opcion_local, n_clusters_max, 0)
        return criterio_distancia, descripcion_opcion, opcion_local, wtgs_por_set_filtrada

    # ------------------------------------------------
    # Menú de opciones
    # ------------------------------------------------
    def menu_opciones():
        #print('Opción 1: Encuentra el óptimo con criterio de distancia')
        #print('Opción 2: Coordenadas de SETs dadas.')
        #print('Opción 3: Encuentra el óptimo de SETs modulares de ', max_trafo_power/10**6, ' MVA')
        #print('Opción 4: Coordenadas de SETs dadas. SETs modulares de ', max_trafo_power/10**6, ' MVA.')
        #print('Opción 5: Todas las opciones + Comparativa')
        opcion_local = opcion
        return opcion_local

    opcion = menu_opciones()

    # ------------------------------------------------
    # Ejecución por opción
    # ------------------------------------------------
    if opcion == 1:
        # *** SIN barrido: usa EXACTAMENTE n_clusters_max ***
        wtg_fuera_radio = 0
        descripcion_opcion = 'SET calculadas. Criterio 10.5km'
        i = int(n_clusters_max)
        if i < 1:
            i = 1

        # Ejecutar Opción 1 SOLO con i = n_clusters_max
        criterio_distancia, descripcion_opcion, opcion, resumen_sets_funcion_op1, wtg_fuera_radio, wtgs_por_set_filtrada_op1 = opcion1(descripcion_opcion, opcion, i)

        base_excel_op1 = f"{descripcion_opcion} {i} SETs, opción {opcion}.xlsx"
        archivos_a_procesar = [os.path.join('WTGs fuera de radio', base_excel_op1)] if wtg_fuera_radio == 1 else [base_excel_op1]

        resumen_op1 = procesar_excels(
            wtg_opcion1=wtg_fuera_radio, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op1,
            archivo_final=os.path.join('RESULTADOS', f"resultados_circuitos_opcion1_{i}_SETs.xlsx"),
            archivo_perdidas=os.path.join('RESULTADOS', f"Pérdidas_Totales_opcion1_{i}_SETs.xlsx"),
            archivos_filtrar=archivos_a_procesar
        )


    elif opcion == 2:
        wtg_fuera_radio = 0
        descripcion_opcion = 'Coordenadas de SETs dadas. Criterio 10.5km'
        criterio_distancia, resumen_sets_funcion_op2, descripcion_opcion, opcion, wtgs_por_set_filtrada_op2 = opcion2(descripcion_opcion, opcion)

        base_excel_op2 = f"{descripcion_opcion}, opción {opcion}.xlsx"
        resumen_op2 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op2,
            archivo_final=os.path.join('RESULTADOS', 'resumen_circuitos_opcion2.xlsx'),
            archivo_perdidas=os.path.join('RESULTADOS', 'Pérdidas_Totales_opcion2.xlsx'),
            archivos_filtrar=[base_excel_op2]
        )



    elif opcion == 3:
        wtg_fuera_radio = 0
        descripcion_opcion = 'SET calculadas. Criterio límite de corriente en barras (SETs modulares pequeñas)'
        descripcion_opcion, opcion, wtgs_por_set_filtrada_op3 = opcion3(descripcion_opcion, opcion)

        sets_coord_op3 = np.array([[np.mean([p[1] for p in grupo]), np.mean([p[2] for p in grupo])] for grupo in wtgs_por_set_filtrada_op3])
        resumen_sets_funcion_op3 = calculos_comunes(sets_coord_op3, descripcion_opcion, 0, wtgs_por_set_filtrada_op3, opcion, n_clusters_max, 0)

        base_excel_op3 = f"{descripcion_opcion}, opción {opcion}.xlsx"
        resumen_op3 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op3,
            archivo_final=os.path.join('RESULTADOS', 'resumen_circuitos_opcion3.xlsx'),
            archivo_perdidas=os.path.join('RESULTADOS', 'Pérdidas_Totales_opcion3.xlsx'),
            archivos_filtrar=[base_excel_op3]
        )

    elif opcion == 4:
        wtg_fuera_radio = 0
        descripcion_opcion = 'Coordenadas de SETs dadas. Criterio límite de corriente en barras (SETs modulares pequeñas)'
        criterio_distancia, descripcion_opcion, opcion, wtgs_por_set_filtrada_op4 = opcion4(descripcion_opcion, opcion)

        sets_coord_op4 = np.array([[np.mean([p[1] for p in grupo]), np.mean([p[2] for p in grupo])] for grupo in wtgs_por_set_filtrada_op4])
        resumen_sets_funcion_op4 = calculos_comunes(sets_coord_op4, descripcion_opcion, 0, wtgs_por_set_filtrada_op4, opcion, n_clusters_max, 0)

        base_excel_op4 = f"{descripcion_opcion}, opción {opcion}.xlsx"
        resumen_op4 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op4,
            archivo_final=os.path.join('RESULTADOS', 'resumen_circuitos_opcion4.xlsx'),
            archivo_perdidas=os.path.join('RESULTADOS', 'Pérdidas_Totales_opcion4.xlsx'),
            archivos_filtrar=[base_excel_op4]
        )

    elif opcion == 5:
        # Recopilatorio global (mantengo el barrido aquí intencionalmente)
        resumen_global = []
        detalles_circuitos_dfs = []

        # Opción 1 (barrido 1..n_clusters_max)
        opcion_local = 1
        descripcion_opcion_local = 'SET calculadas. Criterio 10.5km'
        for i in range(1, n_clusters_max + 1):
            criterio_distancia, descripcion_opcion_local, opcion_local, resumen_sets_funcion_op1, wtg_fuera_radio, wtgs_por_set_filtrada_op1 = opcion1(descripcion_opcion_local, opcion_local, i)

            base_excel_op1 = f"{descripcion_opcion_local} {i} SETs, opción {opcion_local}.xlsx"
            archivos_a_procesar = [os.path.join('WTGs fuera de radio', base_excel_op1)] if wtg_fuera_radio == 1 else [base_excel_op1]
            resumen_op1 = procesar_excels(
                wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
                carpeta_resultados='RESULTADOS',
                resumen_sets_funcion=resumen_sets_funcion_op1,
                archivo_final=os.path.join('RESULTADOS', f"resultados_circuitos_opcion1_{i}_SETs.xlsx"),
                archivo_perdidas=os.path.join('RESULTADOS', f"Pérdidas_Totales_opcion1_{i}_SETs.xlsx"),
                archivos_filtrar=archivos_a_procesar
            )
            for fila in resumen_op1["resumen_por_config"]:
                fila_rec = fila.copy()
                fila_rec["Opción"] = f"1 (SETs={i})"
                resumen_global.append(fila_rec)
            detalles_circuitos_dfs.append(resumen_op1["detalle_por_set"])

        # Opción 2
        opcion_local = 2
        descripcion_opcion_local = 'Coordenadas de SETs dadas. Criterio 10.5km'
        criterio_distancia, resumen_sets_funcion_op2, descripcion_opcion_local, opcion_local, wtgs_por_set_filtrada_op2 = opcion2(descripcion_opcion_local, opcion_local)
        base_excel_op2 = f"{descripcion_opcion_local}, opción {opcion_local}.xlsx"
        resumen_op2 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op2,
            archivo_final=os.path.join('RESULTADOS', 'resumen_circuitos_opcion2.xlsx'),
            archivo_perdidas=os.path.join('RESULTADOS', 'Pérdidas_Totales_opcion2.xlsx'),
            archivos_filtrar=[base_excel_op2]
        )
        for fila in resumen_op2["resumen_por_config"]:
            fila_rec = fila.copy()
            fila_rec["Opción"] = "2"
            resumen_global.append(fila_rec)
        detalles_circuitos_dfs.append(resumen_op2["detalle_por_set"])

        # Opción 3
        opcion_local = 3
        descripcion_opcion_local = 'SET calculadas. Criterio límite de corriente en barras (SETs modulares pequeñas)'
        descripcion_opcion_local, opcion_local, wtgs_por_set_filtrada_op3 = opcion3(descripcion_opcion_local, opcion_local)
        sets_coord_op3 = np.array([[np.mean([p[1] for p in grupo]), np.mean([p[2] for p in grupo])] for grupo in wtgs_por_set_filtrada_op3])
        resumen_sets_funcion_op3 = calculos_comunes(sets_coord_op3, descripcion_opcion_local, 0, wtgs_por_set_filtrada_op3, opcion_local, n_clusters_max, 0)
        base_excel_op3 = f"{descripcion_opcion_local}, opción {opcion_local}.xlsx"
        resumen_op3 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op3,
            archivo_final=os.path.join('RESULTADOS', 'resumen_circuitos_opcion3.xlsx'),
            archivo_perdidas=os.path.join('RESULTADOS', 'Pérdidas_Totales_opcion3.xlsx'),
            archivos_filtrar=[base_excel_op3]
        )
        for fila in resumen_op3["resumen_por_config"]:
            fila_rec = fila.copy()
            fila_rec["Opción"] = "3"
            resumen_global.append(fila_rec)
        detalles_circuitos_dfs.append(resumen_op3["detalle_por_set"])

        # Opción 4
        opcion_local = 4
        descripcion_opcion_local = 'Coordenadas de SETs dadas. Criterio límite de corriente en barras (SETs modulares pequeñas)'
        criterio_distancia, descripcion_opcion_local, opcion_local, wtgs_por_set_filtrada_op4 = opcion4(descripcion_opcion_local, opcion_local)
        sets_coord_op4 = np.array([[np.mean([p[1] for p in grupo]), np.mean([p[2] for p in grupo])] for grupo in wtgs_por_set_filtrada_op4])
        resumen_sets_funcion_op4 = calculos_comunes(sets_coord_op4, descripcion_opcion_local, 0, wtgs_por_set_filtrada_op4, opcion_local, n_clusters_max, 0)
        base_excel_op4 = f"{descripcion_opcion_local}, opción {opcion_local}.xlsx"
        resumen_op4 = procesar_excels(
            wtg_opcion1=0, wtg_opcion2=0, wtg_opcion3=0, wtg_opcion4=0,
            carpeta_resultados='RESULTADOS',
            resumen_sets_funcion=resumen_sets_funcion_op4,
            archivo_final=os.path.join('RESULTADOS', 'resumen_circuitos_opcion4.xlsx'),
            archivo_perdidas=os.path.join('RESULTADOS', 'Pérdidas_Totales_opcion4.xlsx'),
            archivos_filtrar=[base_excel_op4]
        )
        for fila in resumen_op4["resumen_por_config"]:
            fila_rec = fila.copy()
            fila_rec["Opción"] = "4"
            resumen_global.append(fila_rec)
        detalles_circuitos_dfs.append(resumen_op4["detalle_por_set"])

        # Excel recopilatorio global
        df_resumen_global = pd.DataFrame(resumen_global)
        cols_orden = ["Opción","Configuración","Pérdidas en kW","Pérdidas en %",
                      "Circuitos 3 WTGs","Circuitos 2 WTGs","Nº Total Circuitos",
                      "Supply 120mm2 (m)","Supply 300mm2 (m)","Supply 630mm2 (m)","Supply earthing (m)"]
        df_resumen_global = df_resumen_global[[c for c in cols_orden if c in df_resumen_global.columns]]

        df_detalle_circuitos = pd.concat(detalles_circuitos_dfs, axis=1) if detalles_circuitos_dfs else pd.DataFrame()

        out_path = os.path.join("RECOPILATORIO", "RECOPILATORIO_Perdidas_y_Circuitos.xlsx")
        with pd.ExcelWriter(out_path, engine="openpyxl") as writer:
            df_resumen_global.to_excel(writer, sheet_name="Resumen Global", index=False)
            df_detalle_circuitos.to_excel(writer, sheet_name="Detalle Circuitos", index=False)
        print(f"Archivo '{out_path}' generado con éxito.")

        # Comparativa CAPEX
        folder_path = 'RESULTADOS'
        resumen = []
        for filename in os.listdir(folder_path):
            if filename.endswith('.xlsx'):
                file_path = os.path.join(folder_path, filename)
                try:
                    df_capex = pd.read_excel(file_path, engine='openpyxl')
                    if 'CAPEX' in df_capex.columns:
                        capex_values = df_capex['CAPEX'].apply(lambda x: float(str(x).replace(',', '')) if pd.notnull(x) else None).dropna().tolist()
                        capex_value = capex_values[0] if capex_values else None
                    else:
                        capex_value = None
                except Exception:
                    capex_value = None
                resumen.append({'Archivo': filename, 'CAPEX': capex_value})
        df_resumen_capex = pd.DataFrame(resumen)
        output_file = os.path.join('RESULTADOS', 'CAPEX_resumen_resultados.xlsx')
        df_resumen_capex.to_excel(output_file, index=False, engine='openpyxl')
        print(f"Archivo '{output_file}' generado.")

    return resumen_sets_funcion_op1,sets_coord_global
#------prueba
